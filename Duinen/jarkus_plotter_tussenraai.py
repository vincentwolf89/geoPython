import xlrd
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from itertools import groupby
import math
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from mpltools import color
from os import path, listdir
from os.path import splitext
from PIL import Image
from glob import glob
import cv2
import os





def roundup(x):

    return int(math.ceil(x / 100.0)) * 100

def rounddown(x):

    return int(math.floor(x / 100.0)) * 100



excel = r"C:\Users\vince\Dropbox\Wolfwater\HHNK\plots\extra_jrk.xlsx"
adds = r"C:\Users\vince\Dropbox\Wolfwater\HHNK\plots\invoegen_extra.xlsx"

outpath = "C:/Users/vince/Desktop/rest_13_3"



def load_transects():
    global list_raai, list_jaar, list_x, list_z

    book = openpyxl.load_workbook(excel)
    sheet = book.active
    # set row with raainummers
    first_row = sheet[3]
    # define lists
    list_raai = []
    list_jaar = []
    list_x = []
    list_z = []


    for item in first_row:
        value = item.value

        if type(value) == long:

            loc = coordinate_from_string(item.coordinate)
            kolom_z = column_index_from_string(loc[0])
            kolom_x = (get_column_letter(kolom_z-1))
            kolom_z_ = (get_column_letter(kolom_z))
            jaar = sheet.cell(row=4, column=kolom_z).value



            values_x = [c.value for c in sheet[kolom_x][8:]]
            values_z = [c.value for c in sheet[kolom_z_][8:]]

            for item in values_x:
                list_x.append(item)

            for item in values_z:
                list_z.append(item)

                list_raai.append(value)
                list_jaar.append(jaar)



        else:
            continue


    print "import klaar"


def load_adds():
    global raaien, waterstanden, waterstanden_2, maatgevend_jaar, golfhoogtes, golfperiodes, golfhoogtes_2, golfperiodes_2

    book = openpyxl.load_workbook(adds)
    sheet = book.active

    raaien = []
    waterstanden = []
    waterstanden_2 = []
    maatgevend_jaar =[]
    golfhoogtes = []
    golfperiodes = []
    golfhoogtes_2 =[]
    golfperiodes_2 = []

    values_raai = [c.value for c in sheet['A'][1:]]
    values_waterstand = [c.value for c in sheet['B'][1:]]
    values_waterstand_2 = [c.value for c in sheet['C'][1:]]
    maatgevend_kolom = [c.value for c in sheet['D'][1:]]
    golfhoogte_kolom = [c.value for c in sheet['E'][1:]]
    golfperiode_kolom = [c.value for c in sheet['G'][1:]]
    golfhoogte_2_kolom = [c.value for c in sheet['F'][1:]]
    golfperiode_2_kolom = [c.value for c in sheet['H'][1:]]

    for raai in values_raai:
        raaien.append(raai)
    for waterstand in values_waterstand:
        waterstanden.append(waterstand)
    for waterstand in values_waterstand_2:
        waterstanden_2.append(waterstand)
    for maatgevend in maatgevend_kolom:
        maatgevend_jaar.append(maatgevend)
    for golfhoogte in golfhoogte_kolom:
        golfhoogtes.append(golfhoogte)
    for golfhoogte_2 in golfhoogte_2_kolom:
        golfhoogtes_2.append(golfhoogte_2)
    for golfperiode in golfperiode_kolom:
        golfperiodes.append(golfperiode)
    for golfperiode_2 in golfperiode_2_kolom:
        golfperiodes_2.append(golfperiode_2)

def plotter():
    plt.style.use('seaborn-whitegrid')

    # df
    df = pd.DataFrame(list(zip(list_raai, list_jaar, list_x, list_z)))
    df = df.astype(float)
    df.columns = ["locatie", "jaar", "x", "z"]
    df.sort_values(['locatie', 'jaar'], ascending=[True, True])

    # df2
    df2 = pd.DataFrame(list(zip(raaien, waterstanden, waterstanden_2, maatgevend_jaar, golfhoogtes, golfhoogtes_2, golfperiodes, golfperiodes_2)))
    df2.columns = ["raainummer", "waterstand", "waterstand_2", "jaar", "golfhoogte", "golfhoogte_2","golfperiode","golfperiode_2"]


    grouped = df.groupby(["locatie"])
    jaartallen = range(2004, 2018)

    for name, group in grouped:


        xmax = max(group['x'])
        xmin = min(group['x'])

        if xmax > 1600:
            xmax = 1600
        print xmax, xmin


        #n_lines = 15
        #color.cycle_cmap(n_lines, cmap=plt.cm.binary)
        fig, ax = plt.subplots(figsize=(30, 5))

        # get waterstanden en maatgevende jaren
        if int(name) in raaien:
            row = df2.loc[df2['raainummer'] == (int(name))]
            waterstand = row['waterstand'].item()
            waterstand_2 = row['waterstand_2'].item()
            jaar = row['jaar'].item()
            golfh = row['golfhoogte'].item()
            golfp = row['golfperiode'].item()
            golfh_2 = row['golfhoogte_2'].item()
            golfp_2 = row['golfperiode_2'].item()



        for key, grp in group.groupby(['jaar']):

            x = grp['x']
            y = grp['z']

            ax.plot(x, y, linewidth=1, color='black', label=str(int(key))+" (AHN3 en JARKUS-grid)")
            # sleutel = int(key)
            # if sleutel in jaartallen:
            #     if sleutel == jaar:
            #         del sleutel
            #     else:
            #         if int(name) in raaien:
            #             ax.plot(x, y, linewidth = 2, color='black', label = sleutel)

            # if int(key) == jaar:
            #     ax.plot(x,y,linewidth=3, color='red', zorder = 1, label = str(int(key))+" (Maatgevend)")
            #     ax.legend()


        # lim = ax.get_xlim()
        # lim2 = ax.get_ylim()
        #
        #
        # ymin = lim2[0]



        line1 = ax.hlines(waterstand, xmin, xmax, color='blue', zorder=16, linewidth = 1, label='Waterstand categorie IIv')
        line2 = ax.hlines(waterstand_2, xmin, xmax, color='blue', zorder=17, linewidth=0.7, linestyles=[(0,(3,5))], label='Waterstand categorie Iv')


        ax.set_xlabel('Afstand t.o.v RSP 0 [m]')
        ax.set_ylabel('Hoogte t.o.v. NAP [m]')
        profiel = (int(name))/10
        ax.set_title('Tussenraai '+str(profiel)+' (2016)')
        ax.margins(x=0, y=0.1)

        ax.set_xlim(xmin, xmax)  # TEST

        leg = ax.legend(loc=7, frameon=1)
        frame = leg.get_frame()
        frame.set_facecolor('white')
        leg.set_title('Tussenraai', prop={'size': 12, 'weight': 'heavy'})
        leg._legend_box.align = "left"
        leg.set_zorder(18)

        # plot text

        waterstand_text = (str(waterstand).replace('.', ','))
        waterstand2_text = (str(waterstand_2).replace('.', ','))
        golfh_text = (str(golfh).replace('.', ','))
        golfp_text = (str(golfp).replace('.', ','))
        golfh_2_text = (str(golfh_2).replace('.', ','))
        golfp_2_text = (str(golfp_2).replace('.', ','))


        #t = 'Categorie IIv'+'\nWaterstand categorie IIv: '+waterstand_text+' m NAP' + '\nWaterstand categorie Iv: '+waterstand2_text+' m NAP'+'\nGolfhoogte: '+golfh_text+' m'+'\nGolfperiode: '+golfp_text+' s'
        t = 'Categorie IIv' + '\nWaterstand: NAP + ' + waterstand_text + ' meter' +'\nGolfhoogte: ' + golfh_text + ' m' + '\nGolfperiode: ' + golfp_text + ' s'
        t2 = 'Categorie Iv' + '\nWaterstand: NAP + ' + waterstand2_text + ' meter' +'\nGolfhoogte: ' + golfh_2_text + ' m' + '\nGolfperiode: ' + golfp_2_text + ' s'


        plot1 = ax.text(0.01,0.02, t, size=10,
                 va="bottom", ha="left", multialignment="left", transform=ax.transAxes)




        plot2 = plt.text(0.10,0.02, t2, size=10,va="bottom", ha="left", multialignment="left", transform=ax.transAxes)


        # plot2 = plt.text(xmin +(1.6*(width)), ymin, t2, size=10,
        #                  #va="bottom", ha="left", multialignment="left")

        # plot2 = ax.text(xmin + (2 * (width)), ymin, t2, size=10,
        #                  va="bottom", ha="left", multialignment="left")

        plot3 = ax.text(0.01, 0.18, "Hydraulische belasting", size=11, weight='bold', va="bottom", ha="left",
                 multialignment="left", transform=ax.transAxes)


        major_ticks = np.arange(roundup(xmin), rounddown(xmax), 100)
        #minor_ticks = np.arange(roundup(xmin), rounddown(lim[1]), 100)

        ax.set_xticks(major_ticks)
        #ax.set_xticks(minor_ticks, minor=True)



        # plt.show()
        plt.savefig(path.join(outpath, "tussenraai_{0}.png".format(profiel)), pad_inches=0.02, dpi=300, bbox_inches='tight')
        del fig, ax, xmax, xmin, row, waterstand,waterstand_2,jaar,golfh,golfp,golfh_2,golfp_2,x,y,line1,line2,profiel,leg,frame,waterstand_text,waterstand2_text,golfh_text,golfh_2_text,golfp_text,golfp_2_text,t,t2,plot1,plot2,plot3,major_ticks
        print "grafiek voor profiel " + str(name) + " gemaakt"

        plt.clf()
        plt.cla()
        plt.close()






def png_converter():

    pngs = glob(outpath+'/*.png')

    for j in pngs:
        img = cv2.imread(j)
        cv2.imwrite(j[:-3] + 'tiff', img)
        os.remove(j)


load_adds()
load_transects()
plotter()
# # png_converter()