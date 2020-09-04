import gc
gc.collect()
import arcpy
import math
from arcpy.sa import *
import xlwt
import pandas as pd
from itertools import groupby
from xlsxwriter.workbook import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
# uitzetten melding pandas
pd.set_option('mode.chained_assignment', None)




def average(lijst):
    return sum(lijst) / len(lijst)







def CopyParallelL(plyP,sLength): #functie voor profielen maken haaks op trajectlijn
    part=plyP.getPart(0)
    lArray=arcpy.Array()
    for ptX in part:
        dL=plyP.measureOnLine(ptX)
        ptX0=plyP.positionAlongLine (dL-0.01).firstPoint
        ptX1=plyP.positionAlongLine (dL+0.01).firstPoint
        dX=float(ptX1.X)-float(ptX0.X)
        dY=float(ptX1.Y)-float(ptX0.Y)
        lenV=math.hypot(dX,dY)
        sX=-dY*sLength/lenV;sY=dX*sLength/lenV
        leftP=arcpy.Point(ptX.X+sX,ptX.Y+sY)
        lArray.add(leftP)
    array = arcpy.Array([lArray])
    section=arcpy.Polyline(array)
    return section

def CopyParallelR(plyP,sLength): #functie voor profielen maken haaks op trajectlijn
    part=plyP.getPart(0)
    rArray=arcpy.Array()
    for ptX in part:
        dL=plyP.measureOnLine(ptX)
        ptX0=plyP.positionAlongLine (dL-0.01).firstPoint
        ptX1=plyP.positionAlongLine (dL+0.01).firstPoint
        dX=float(ptX1.X)-float(ptX0.X)
        dY=float(ptX1.Y)-float(ptX0.Y)
        lenV=math.hypot(dX,dY)
        sX=-dY*sLength/lenV;sY=dX*sLength/lenV
        rightP=arcpy.Point(ptX.X-sX, ptX.Y-sY)
        rArray.add(rightP)
    array = arcpy.Array([rArray])
    section=arcpy.Polyline(array)
    return section

def copy_trajectory_lr(trajectlijn,code,afstand):
    existing_fields = arcpy.ListFields(trajectlijn)
    needed_fields = ['OBJECTID','Shape','Shape_Length','SHAPE', 'SHAPE_Length',code]
    for field in existing_fields:
        if field.name not in needed_fields:
            arcpy.DeleteField_management(trajectlijn, field.name)

    arcpy.AddField_management(trajectlijn, "Width", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management(trajectlijn, "Width", afstand, "PYTHON")

    arcpy.CopyFeatures_management(trajectlijn, "river")
    arcpy.CopyFeatures_management(trajectlijn, "land")
    land = "land"
    river = "river"


    with arcpy.da.UpdateCursor(land, ("Shape@", "Width")) as cursor:
        for shp, w in cursor:
            LeftLine = CopyParallelL(shp, w)
            cursor.updateRow((LeftLine, w))

    with arcpy.da.UpdateCursor(river, ("Shape@", "Width")) as cursor:
        for shp, w in cursor:
            RightLine = CopyParallelR(shp, w)
            cursor.updateRow((RightLine, w))
    # print "Water- en landdelen gemaakt"

def set_measurements_trajectory(profielen,trajectlijn,code,stapgrootte_punten,toetspeil): #rechts = rivier, profielen van binnen naar buiten
    # clean feature
    existing_fields = arcpy.ListFields(profielen)
    needed_fields = ['OBJECTID', 'SHAPE', 'SHAPE_Length','Shape','Shape_Length',code,toetspeil,'profielnummer']
    for field in existing_fields:
        if field.name not in needed_fields:
            arcpy.DeleteField_management(profielen, field.name)

    # add needed fields
    #arcpy.AddField_management(profielen, "profielnummer", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management(profielen, "van", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management(profielen, "tot", "DOUBLE", 2, field_is_nullable="NULLABLE")

    #arcpy.CalculateField_management(profielen, "profielnummer", '!OBJECTID!', "PYTHON")

    # split profiles
    rivierlijn = "river"
    landlijn = "land"
    clusterTolerance = 0
    invoer = [profielen, trajectlijn]
    uitvoer = 'snijpunten_centerline'
    arcpy.Intersect_analysis(invoer, uitvoer, "", clusterTolerance, "point")
    arcpy.SplitLineAtPoint_management(profielen, uitvoer, 'profielsplits', 1)

    velden = ['profielnummer', 'van', 'tot',code]

    fieldmappings = arcpy.FieldMappings()
    fieldmappings.addTable('profielsplits')
    fieldmappings.addTable(rivierlijn)
    fieldmappings.addTable(landlijn)
    keepers = velden

    # join splits to river/land parts
    for field in fieldmappings.fields:
        if field.name not in keepers:
            fieldmappings.removeFieldMap(fieldmappings.findFieldMapIndex(field.name))

    arcpy.SpatialJoin_analysis('profielsplits', rivierlijn, 'profieldeel_rivier', "JOIN_ONE_TO_ONE", "KEEP_COMMON", fieldmappings,
                               match_option="INTERSECT")
    arcpy.SpatialJoin_analysis('profielsplits', landlijn, 'profieldeel_land', "JOIN_ONE_TO_ONE", "KEEP_COMMON",
                               fieldmappings,
                               match_option="INTERSECT")

    # create routes
    arcpy.CalculateField_management("profieldeel_rivier", "tot", '!Shape_Length!', "PYTHON")
    arcpy.CalculateField_management("profieldeel_land", "tot", '!Shape_Length!', "PYTHON")
    arcpy.CalculateField_management("profieldeel_rivier", "van", 0, "PYTHON")
    arcpy.CalculateField_management("profieldeel_land", "van", 0, "PYTHON")


    arcpy.CreateRoutes_lr('profieldeel_rivier', "profielnummer", "routes_rivier_", "TWO_FIELDS", "van", "tot", "", "1", "0",
                          "IGNORE", "INDEX")

    arcpy.CreateRoutes_lr('profieldeel_land', "profielnummer", "routes_land_", "TWO_FIELDS", "tot", "van", "", "1",
                          "0", "IGNORE", "INDEX")

    #join code
    velden = ['profielnummer',code]
    fieldmappings = arcpy.FieldMappings()
    fieldmappings.addTable('routes_land_')
    fieldmappings.addTable('routes_rivier_')
    fieldmappings.addTable(trajectlijn)

    keepers = velden
    for field in fieldmappings.fields:
        if field.name not in keepers:
            fieldmappings.removeFieldMap(fieldmappings.findFieldMapIndex(field.name))

    arcpy.SpatialJoin_analysis('routes_rivier_', trajectlijn, 'routes_rivier', "JOIN_ONE_TO_ONE", "KEEP_COMMON",
                               fieldmappings,
                               match_option="INTERSECT")
    arcpy.SpatialJoin_analysis('routes_land_', trajectlijn, 'routes_land', "JOIN_ONE_TO_ONE", "KEEP_COMMON",
                               fieldmappings,
                               match_option="INTERSECT")

    # generate points
    arcpy.GeneratePointsAlongLines_management('routes_land', 'punten_land', 'DISTANCE', Distance= stapgrootte_punten)
    arcpy.GeneratePointsAlongLines_management('routes_rivier', 'punten_rivier', 'DISTANCE', Distance=stapgrootte_punten)


    # id field for joining table
    arcpy.AddField_management('punten_land', 'punt_id', "DOUBLE", field_precision=2, field_is_nullable="NULLABLE")
    arcpy.AddField_management('punten_rivier', 'punt_id', "DOUBLE", field_precision=2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management("punten_land", "punt_id", '!OBJECTID!', "PYTHON")
    arcpy.CalculateField_management("punten_rivier", "punt_id", '!OBJECTID!', "PYTHON")


    # find points along routes
    Output_Event_Table_Properties = "RID POINT MEAS"
    arcpy.LocateFeaturesAlongRoutes_lr('punten_land', 'routes_land', "profielnummer", "1 Meters",
                                       'uitvoer_tabel_land', Output_Event_Table_Properties, "FIRST", "DISTANCE", "ZERO",
                                       "FIELDS", "M_DIRECTON")
    arcpy.LocateFeaturesAlongRoutes_lr('punten_rivier', 'routes_rivier', "profielnummer", "1 Meters",
                                       'uitvoer_tabel_rivier', Output_Event_Table_Properties, "FIRST", "DISTANCE", "ZERO",
                                       "FIELDS", "M_DIRECTON")

    # join fields from table
    arcpy.JoinField_management('punten_land', 'punt_id', 'uitvoer_tabel_land', 'punt_id', 'MEAS')
    arcpy.JoinField_management('punten_rivier', 'punt_id', 'uitvoer_tabel_rivier', 'punt_id', 'MEAS')
    arcpy.AlterField_management('punten_land', 'MEAS', 'afstand')
    arcpy.AlterField_management('punten_rivier', 'MEAS', 'afstand')

    with arcpy.da.UpdateCursor('punten_rivier', ['profielnummer', 'afstand']) as cursor:
        for row in cursor:
            row[1] = row[1]*-1
            cursor.updateRow(row)

    fieldmappings = arcpy.FieldMappings()
    fieldmappings.addTable('punten_land')
    fieldmappings.addTable('punten_rivier')
    fieldmappings.addTable('snijpunten_centerline')

    velden = ['profielnummer', 'afstand', code]
    keepers = velden

    for field in fieldmappings.fields:
        if field.name not in keepers:
            fieldmappings.removeFieldMap(fieldmappings.findFieldMapIndex(field.name))

    arcpy.FeatureToPoint_management("snijpunten_centerline", "punten_centerline")
    arcpy.Merge_management(['punten_land', 'punten_rivier','punten_centerline'], 'punten_profielen', fieldmappings)

    arcpy.CalculateField_management("punten_profielen", "afstand", 'round(!afstand!, 1)', "PYTHON")

    # set centerline values to 0
    with arcpy.da.UpdateCursor('punten_profielen', ['afstand']) as cursor:
        for row in cursor:
            if row[0] == None:
                row[0] = 0
                cursor.updateRow(row)

    print 'Meetpunten op routes gelokaliseerd'

def extract_z_arcpy(invoerpunten, uitvoerpunten, raster): #

    # Test de ArcGIS Spatial Analyst extension license
    arcpy.CheckOutExtension("Spatial")

    # Koppel z-waardes
    ExtractValuesToPoints(invoerpunten, raster, uitvoerpunten,
                          "INTERPOLATE", "VALUE_ONLY")

    # Pas het veld 'RASTERVALU' aan naar 'z_ahn'
    arcpy.AlterField_management(uitvoerpunten, 'RASTERVALU', 'z_ahn')
    print "Hoogtewaarde aan punten gekoppeld"

def add_xy(uitvoerpunten,code,trajectlijn):

    existing_fields = arcpy.ListFields(uitvoerpunten)
    needed_fields = ['OBJECTID', 'Shape', 'profielnummer', 'afstand', 'z_ahn', code]
    for field in existing_fields:
        if field.name not in needed_fields:
            arcpy.DeleteField_management(trajectlijn, field.name)

    arcpy.env.outputCoordinateSystem = arcpy.Describe(uitvoerpunten).spatialReference
    # Set local variables
    in_features = uitvoerpunten
    properties = "POINT_X_Y_Z_M"
    length_unit = ""
    area_unit = ""
    coordinate_system = ""

    # Generate the extent coordinates using Add Geometry Properties tool
    arcpy.AddGeometryAttributes_management(in_features, properties, length_unit,
                                           area_unit,
                                           coordinate_system)

    arcpy.AlterField_management(uitvoerpunten, 'POINT_X', 'x')
    arcpy.AlterField_management(uitvoerpunten, 'POINT_Y', 'y')

    print "XY-coordinaten aan punten gekoppeld"


def create_contour_lines(trajectlijn,raster,contourlines,distance):
    # buffer trajectlijn
    buffer_traject = 'buffer_traject'

    print "buffer gemaakt"
    arcpy.Buffer_analysis(trajectlijn,buffer_traject, "50 Meters", "FULL", "ROUND", "NONE", "", "PLANAR")

    # clip raster with buffer
    arcpy.Clip_management(raster, "110891,550900001 435080,382600008 140346,480208685 445543,559100002",
                          "clip_dijktraject", buffer_traject, "-3,402823e+038", "ClippingGeometry",
                          "MAINTAIN_EXTENT")

    print "raster geclipt"
    arcpy.Contour_3d("clip_dijktraject", contourlines, distance, "0", "1")
    print "contour gemaakt"

def snap_to_contour(puntenset,contour,output_tabel):
    # neartable
    arcpy.GenerateNearTable_analysis(puntenset, contour, output_tabel, "5 Meters", "NO_LOCATION", "ANGLE", "CLOSEST", "0", "PLANAR")

    arcpy.CopyFeatures_management(contour, "contour_copy")

    # delete onnodig
    id_punten =[]
    id_contour = []
    with arcpy.da.UpdateCursor(output_tabel, ("IN_FID", "NEAR_FID")) as cursor:
        for row in cursor:
            id_punten.append(row[0])
            id_contour.append(row[1])
    contour_copy = 'contour_copy'
    with arcpy.da.UpdateCursor(contour_copy, ("OBJECTID")) as cursor:
        for row in cursor:
            if row[0] not in id_contour:
                cursor.deleteRow()
            else:
                pass

    # snap
    arcpy.Snap_edit(puntenset, "contour_copy EDGE '3 Meters'")
    arcpy.CopyFeatures_management(puntenset, puntenset+"_snap")


def kruinhoogte_groepen(uitvoerpunten,stapgrootte_punten,afronding,code):

    # Verwijder punten zonder z_waarde en rond afstand af
    with arcpy.da.UpdateCursor(uitvoerpunten, ['afstand', 'z_ahn'],sql_clause = (None,"ORDER BY afstand ASC")) as cursor:
        for row in cursor:
            if row[1] is None:
                cursor.deleteRow()
            else:
                row[0] = round(row[0],afronding)
                cursor.updateRow(row)
    del cursor

    existing_fields = arcpy.ListFields(uitvoerpunten)
    needed_fields = ['OBJECTID', 'SHAPE', 'SHAPE_Length', 'Shape', 'Shape_Length','afstand','z_ahn','profielnummer', code,'x','y']
    for field in existing_fields:
        if field.name not in needed_fields:
            arcpy.DeleteField_management(uitvoerpunten, field.name)

    # add needed fields
    arcpy.AddField_management(uitvoerpunten, "groep", "DOUBLE", 2, field_is_nullable="NULLABLE")

    # Sorteer punten op profielnummer, afstand
    sort_fields = [["profielnummer", "ASCENDING"], ["afstand", "ASCENDING"]]
    arcpy.Sort_management(uitvoerpunten, "sorted", sort_fields)

    # Bepaal groepsnummer
    with arcpy.da.UpdateCursor('sorted', ['afstand', 'z_ahn', 'groep', 'profielnummer']) as cursor:
            for k, g in groupby(cursor, lambda x: x[3]):
                p = g.next()[0] # eerste waarde

                value = 1

                for row in g:
                    c = row[0] # volgende waarde
                    if abs(c-p) <= stapgrootte_punten:
                        row[2] = value
                        cursor.updateRow(row)
                    else:
                        value+=1
                        # print "Nieuwe volumegroep gemaakt op profiel, g"
                        # pass

                    p = row[0]

                    row[2] = value
                    cursor.updateRow(row)

    del cursor

    # Aanpassen eerste groepwaardes (van None naar 1)
    with arcpy.da.UpdateCursor('sorted', ['groep']) as cursor:
        for row in cursor:
            if row[0] == None:
                row[0] = 1
                cursor.updateRow(row)
    del cursor
    # Vervang niet-gesorteerede puntenset
    arcpy.CopyFeatures_management("sorted", uitvoerpunten)
    print 'Kruinhoogte voor alle aaneengesloten groepen binnen profielen bepaald'



def max_kruinhoogte(uitvoerpunten, profielen, code, uitvoer_maxpunten,min_afstand,max_afstand,toetspeil):
    # feature class to numpy array
    array = arcpy.da.FeatureClassToNumPyArray(uitvoerpunten,
                                              ('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn', 'groep'))
    # dataframe, algemeen
    df = pd.DataFrame(array)
    sorted = df.sort_values(['profielnummer', 'afstand'], ascending=[True, True])

    # dataframe voor maximale kruinhoogtes over 1.5m voor alle profielen
    df_maxwaardes = pd.DataFrame(columns=['profielnummer', 'max_kruinhoogte'])

    # sorteer op profielnummer
    groep_profiel = sorted.groupby('profielnummer')

    # itereer over profielen
    for name, groep in groep_profiel:
        # dataframe voor maximale kruinhoogtes over 1.5 m voor alle groepen per profiel
        df_profiel = pd.DataFrame(columns=['profielnummer', 'max_kruinhoogte'])
        df_profiel_tester = pd.DataFrame(columns=['profielnummer', 'max_kruinhoogte'])
        # itereer over groepen binnen profielen
        groep_punten = groep.groupby(["groep"])

        for groep, groep_binnen_profiel in groep_punten:
            # bepaal gemiddelde kruinhoogte over 1.5 m(3 meetpunten), extra kolom 'max_kr'
            groep_binnen_profiel['max_kr'] = groep_binnen_profiel.iloc[:, 4].rolling(window=3).mean()
            # shift kolom 'max_kr' om middenpunt voor ieder rolling mean te verkrijgen (punt 2 van 3 meetpunten)
            groep_binnen_profiel['middelpunt_max'] = groep_binnen_profiel['max_kr'].shift(-1)

            # zoeken naar middelpunt rolling mean voor maken centerline max_kruinhoogte


            for i, row in groep_binnen_profiel.iterrows():
                afstand = row['afstand']
                if afstand > min_afstand and afstand < max_afstand:
                    df_profiel_tester.loc[row['OBJECTID']] = name, row['middelpunt_max']
        if df_profiel_tester.empty == True:
            pass
        else:
            # oude code, vervangen door nieuwe sectie
            # df_profielmax = df_profiel_tester.max()[1]
            # id_df_profielmax = df_profiel_tester.idxmax()[1]

            df_profielmax = df_profiel_tester['max_kruinhoogte'].max()
            id_df_profielmax = df_profiel_tester['max_kruinhoogte'].idxmax()


            df_profiel.loc[id_df_profielmax] = name, df_profielmax

        # check of profiel een maximale kruinhoogte heeft
        if df_profiel.empty == True:
            pass
        else:
            kr_max = df_profiel['max_kruinhoogte'].max()
            OID_max = df_profiel['max_kruinhoogte'].idxmax()

            df_maxwaardes.loc[OID_max] = name, kr_max
            # print df_profiel

    # print per profiel de gevonden resultaten
    print df_maxwaardes

    # update profielen met maximale kruinhoogte
    existing_fields = arcpy.ListFields(profielen)
    needed_fields = ['OBJECTID', 'SHAPE', 'SHAPE_Length', 'Shape_Length', 'Shape', 'profielnummer', code,toetspeil]
    for field in existing_fields:
        if field.name not in needed_fields:
            arcpy.DeleteField_management(profielen, field.name)

    # voeg veld voor maximale kruinhoogte toe aan lijn-laag profielen
    arcpy.AddField_management(profielen, "maxKruinhoogte2014", "DOUBLE", 2, field_is_nullable="NULLABLE")
    with arcpy.da.UpdateCursor(profielen, ['profielnummer', 'maxKruinhoogte2014']) as cursor:
        for row in cursor:
            profielnummer = row[0]
            for i, row in df_maxwaardes.iterrows():
                if row['profielnummer'] == profielnummer:
                    row[1] = round(row['max_kruinhoogte'], 2)
                    cursor.updateRow(row)

    # genereer puntenlijn met maximale gemiddelde kruinhoogte
    arcpy.CopyFeatures_management(uitvoerpunten, uitvoer_maxpunten)
    list_oid = df_maxwaardes.index.values.tolist()

    # selecteer middelpunten maximale kruinhoogte over 1.5 m
    with arcpy.da.UpdateCursor(uitvoer_maxpunten, ("OBJECTID")) as cursor:
        for row in cursor:
            if row[0] in list_oid:
                pass
            else:
                cursor.deleteRow()

    print 'Maximale kruinhoogte voor ieder profiel bepaald indien mogelijk'
def generate_profiles(profiel_interval,profiel_lengte_land,profiel_lengte_rivier,trajectlijn,code,toetspeil,profielen):
    # traject to points


    arcpy.GeneratePointsAlongLines_management(trajectlijn, 'traject_punten', 'DISTANCE', Distance=profiel_interval, Include_End_Points='END_POINTS')
    arcpy.AddField_management('traject_punten', "profielnummer", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management('traject_punten', "lengte_landzijde", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management('traject_punten', "lengte_rivierzijde", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management('traject_punten', "profielnummer", '!OBJECTID!', "PYTHON")
    arcpy.CalculateField_management('traject_punten', "lengte_landzijde", profiel_lengte_land, "PYTHON")
    arcpy.CalculateField_management('traject_punten', "lengte_rivierzijde", profiel_lengte_rivier, "PYTHON")

    # route voor trajectlijn
    # arcpy.CreateRoutes_lr(trajectlijn, code, "route_traject", "LENGTH", "", "", "UPPER_LEFT", "1", "0", "IGNORE", "INDEX")

    existing_fields = arcpy.ListFields(trajectlijn)
    needed_fields = ['OBJECTID', 'SHAPE', 'SHAPE_Length','Shape','Shape_Length',code,toetspeil]
    for field in existing_fields:
        if field.name not in needed_fields:
            arcpy.DeleteField_management(trajectlijn, field.name)
    arcpy.AddField_management(trajectlijn, "van", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management(trajectlijn, "tot", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management(trajectlijn, "van", 0, "PYTHON")
    # arcpy.CalculateField_management(trajectlijn, "tot", "!Shape_Length!", "PYTHON")
    arcpy.CalculateField_management(trajectlijn, "tot", "round(!shape.length!)", "PYTHON")
    arcpy.CreateRoutes_lr(trajectlijn, code, 'route_traject', "TWO_FIELDS", "van", "tot", "", "1",
                          "0", "IGNORE", "INDEX")


    # locate profielpunten
    arcpy.LocateFeaturesAlongRoutes_lr('traject_punten', 'route_traject', code, "1.5 Meters", 'tabel_traject_punten',
                                       "RID POINT MEAS", "FIRST", "DISTANCE", "ZERO", "FIELDS",
                                       "M_DIRECTON")

    # offset rivierdeel profiel
    arcpy.MakeRouteEventLayer_lr('route_traject', code, 'tabel_traject_punten', "rid POINT meas", 'deel_rivier',
                                 "lengte_rivierzijde", "NO_ERROR_FIELD", "NO_ANGLE_FIELD", "NORMAL", "ANGLE", "RIGHT",
                                 "POINT")

    arcpy.MakeRouteEventLayer_lr('route_traject', code, 'tabel_traject_punten', "rid POINT meas", 'deel_land',
                                 "lengte_landzijde", "NO_ERROR_FIELD", "NO_ANGLE_FIELD", "NORMAL", "ANGLE", "LEFT",
                                 "POINT")
    # temp inzicht layer
    arcpy.CopyFeatures_management('deel_rivier', "temp_rivierdeel")
    arcpy.CopyFeatures_management('deel_land', "temp_landdeel")
    arcpy.AddField_management('temp_rivierdeel', "id", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management('temp_landdeel', "id", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management('temp_rivierdeel', "id", 2, "PYTHON")
    arcpy.CalculateField_management('temp_landdeel', "id", 1, "PYTHON")





    arcpy.Merge_management("'temp_rivierdeel';'temp_landdeel'", 'merge_profielpunten')
    arcpy.PointsToLine_management('merge_profielpunten', profielen, "profielnummer", "id", "NO_CLOSE")

    arcpy.SpatialJoin_analysis(profielen, trajectlijn, 'profielen_temp', "JOIN_ONE_TO_ONE", "KEEP_ALL", match_option="INTERSECT")
    arcpy.CopyFeatures_management('profielen_temp', profielen)
    # arcpy.FlipLine_edit(profielen)

    print 'Profielen gemaakt op trajectlijn'


def locate_profiles(profielen,dijkpalen,split_traject,RID,profileID,profielenUitvoer):
    

    # bepaal overgangen (snijpunten dijkpalen)
    tempOvergangen = "tempOvergangen"
    arcpy.FeatureVerticesToPoints_management(split_traject, tempOvergangen, "BOTH_ENDS")
    
    #### non splits
    arcpy.MakeFeatureLayer_management(profielen, 'templayer') 
    arcpy.SelectLayerByLocation_management('templayer', "INTERSECT", tempOvergangen, "", "NEW_SELECTION", "INVERT")
    arcpy.CopyFeatures_management('templayer', "profielenNonSplit")

    # templayer for intersect points:
    tempIntersect = "tempIntersect"
    arcpy.Intersect_analysis(["profielenNonSplit",split_traject], tempIntersect, "ALL", "", "POINT")

    # locate features along route: "temp_intersect"
    tempLocTable = "tempLocTable"
    arcpy.LocateFeaturesAlongRoutes_lr(tempIntersect, split_traject, RID, "1 Meters", tempLocTable,
    "RID POINT MEAS", "FIRST", "DISTANCE", "ZERO", "FIELDS", "M_DIRECTON")

    # join meas and dp to profiles
    arcpy.JoinField_management("profielenNonSplit", profileID, tempLocTable, profileID, "MEAS;RFTIDENT")

    # add field for locatie
    arcpy.AddField_management("profielenNonSplit", 'metreringDp', "TEXT")
    arcpy.CalculateField_management("profielenNonSplit", "metreringDp", "!RFTIDENT!+\"+\"+str(round( !MEAS!))", "PYTHON", "")

    ####
    # splits
    arcpy.MakeFeatureLayer_management(profielen, 'templayer') 
    arcpy.SelectLayerByLocation_management('templayer', "INTERSECT", tempOvergangen, "", "NEW_SELECTION", "")
    arcpy.CopyFeatures_management('templayer', "profielenSplit")

    arcpy.SpatialJoin_analysis("profielenSplit", dijkpalen, "profielenSplitDp", "#", "#", match_option="CLOSEST")
    
    # add field for locatie
    arcpy.AddField_management("profielenSplitDp", 'metreringDp', "TEXT")
    arcpy.CalculateField_management("profielenSplitDp", "metreringDp", "!RFTIDENT!+\"+0.0\"", "PYTHON", "")


    #### samenvoegen
    arcpy.Merge_management(["profielenSplitDp","profielenNonSplit"], profielenUitvoer)

    veldenNodig = ["OBJECTID","Shape","Shape_Length","profielnummer","metreringDp","dijkvak","dijkpaal"]

    veldenTotaalObject = arcpy.ListFields(profielenUitvoer)
    veldenTotaal = []
            
    for veld in veldenTotaalObject:
        veldenTotaal.append(str(veld.name))

    for veld in veldenTotaal:
        if veld in veldenNodig:
            pass
        else:
            arcpy.DeleteField_management(profielenUitvoer,veld)

    

    #### schoonmaken
    arcpy.Delete_management(tempOvergangen)
    arcpy.Delete_management(tempLocTable)
    arcpy.Delete_management("profielenNonSplit")
    arcpy.Delete_management("profielenSplit")
    arcpy.Delete_management("profielenSplitDp")



    print ("Profielen gekoppeld aan dijkpalen en metrering berekend")

    
    
    



    # arcpy.Delete_management(tempIntersect)
    # arcpy.Delete_management(tempLocTable)






    


def join_mg_profiles(trajectlijn, profielen,profielen_mg,profielen_plus):
    # select from trajectlijn
    arcpy.MakeFeatureLayer_management(profielen_mg, 'templaag_profielen_mg')
    arcpy.MakeFeatureLayer_management(profielen, 'templaag_profielen')
    arcpy.SelectLayerByLocation_management('templaag_profielen_mg', 'intersect', trajectlijn)
    arcpy.CopyFeatures_management('templaag_profielen_mg', 'maatgevend_profiel')
    # merge
    arcpy.Merge_management(['templaag_profielen', 'maatgevend_profiel'], profielen_plus)
    print "maatgevende profielen toegevoegd aan profielset"
def kruinbepalen(invoer, code, uitvoer_binnenkruin, uitvoer_buitenkruin,verschil_maxkruin,min_afstand,max_afstand):
    array = arcpy.da.FeatureClassToNumPyArray(invoer, ('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn'))
    df = pd.DataFrame(array)
    df2 = df.dropna()
    sorted = df2.sort_values(['profielnummer', 'afstand'], ascending=[True, True])
    grouped = sorted.groupby('profielnummer')

    list_id_bik = []
    list_id_buk = []

    afstand_bik = max_afstand
    afstand_buk = min_afstand

    for name, group in grouped:
        # maximale kruinhoogte
        max_kruin = max(group['z_ahn'])

        landzijde = group.sort_values(['afstand'], ascending=False)  # afnemend, landzijde
        rivierzijde = group.sort_values(['afstand'], ascending=True)  # toenemend, rivierzijde

        # bik
        for index, row in landzijde.iterrows():
            if row['z_ahn'] > max_kruin - verschil_maxkruin and row['afstand'] < afstand_bik and row['afstand'] > 0:
                x_bik = row['afstand']
                y_bik = row['z_ahn']
                list_id_bik.append(row['OBJECTID'])
                break

        # buk
        for index, row in rivierzijde.iterrows():
            if row['z_ahn'] > max_kruin - verschil_maxkruin and row['afstand'] > afstand_buk and row['afstand'] < 0:
                x_buk = row['afstand']
                y_buk = row['z_ahn']
                list_id_buk.append(row['OBJECTID'])
                break

    # wegschrijven naar gis

    # binnenkruin
    if list_id_bik:
        arcpy.MakeFeatureLayer_management(invoer, 'punten_binnenkruin_temp')
        punten_bik = arcpy.SelectLayerByAttribute_management('punten_binnenkruin_temp', "ADD_TO_SELECTION",
                                                             "OBJECTID in (" + str(list_id_bik)[1:-1] + ")")
        arcpy.CopyFeatures_management('punten_binnenkruin_temp', uitvoer_binnenkruin)

    # buitenkruin
    if list_id_buk:
        arcpy.MakeFeatureLayer_management(invoer, 'punten_buitenkruin_temp')
        punten_buk = arcpy.SelectLayerByAttribute_management('punten_buitenkruin_temp', "ADD_TO_SELECTION",
                                                             "OBJECTID in (" + str(list_id_buk)[1:-1] + ")")
        arcpy.CopyFeatures_management('punten_buitenkruin_temp', uitvoer_buitenkruin)


    print 'Kruinpunten bepaald'

def excel_writer(uitvoerpunten,code,excel,id,trajecten,toetspeil,min_plot,max_plot):
    # toetshoogte aan uitvoerpunten koppelen
    arcpy.JoinField_management(uitvoerpunten, code, trajecten, code, toetspeil)

    # binnenhalen van dataframe
    if toetspeil == 999:
        array = arcpy.da.FeatureClassToNumPyArray(uitvoerpunten,('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn', 'x', 'y'))
    else:
        array = arcpy.da.FeatureClassToNumPyArray(uitvoerpunten, ('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn', 'x', 'y', toetspeil))

    df = pd.DataFrame(array)
    df = df.dropna()
    sorted = df.sort_values(['profielnummer', 'afstand'], ascending=[True, True])

    # opbouw xlsx
    workbook = Workbook(excel)
    worksheet = workbook.add_worksheet()

    # stijl toevoegen voor headers
    bold = workbook.add_format({'bold': True})


    # schrijf kolomnamen
    worksheet.write(0, 0, "Profielnummer", bold)
    worksheet.write(0, 1, "Afstand [m]", bold)
    worksheet.write(0, 2, "Hoogte AHN3 [m NAP]", bold)
    worksheet.write(0, 3, "x [RD]", bold)
    worksheet.write(0, 4, "y [RD]", bold)

    # schrijf kolommen vanuit df
    worksheet.write_column('A2', sorted['profielnummer'])
    worksheet.write_column('B2', sorted['afstand'])
    worksheet.write_column('C2', sorted['z_ahn'])
    worksheet.write_column('D2', sorted['x'])
    worksheet.write_column('E2', sorted['y'])

    # groepeer per profielnummer
    grouped = sorted.groupby('profielnummer')

    # definieer startrij
    startpunt = 2


    # lege lijngrafiek invoegen met zowel afstand als hoogte als invoer
    line_chart1 = workbook.add_chart({'type': 'scatter',
                                 'subtype': 'straight'})

    ## toetshoogte, aan/uit
    # toetshoogte toevoegen als horizontale lijn, deel 1 voor legenda
    # minimum = min(sorted['afstand'])
    # maximum = max(sorted['afstand'])
    # th = sorted[toetspeil].iloc[0]
    #
    # worksheet.write('K8', minimum)
    # worksheet.write('K9', maximum)
    # worksheet.write('K10', th)
    # worksheet.write('K11', th)


    # line_chart1.add_series({
    #     'name': 'toetshoogte: ' + str(th) + ' m NAP',
    #
    #     'categories': '=Sheet1!$K$8:$K$9',
    #     'values': '=Sheet1!$K$10:$K$11',
    #     'line': {
    #         'color': 'red',
    #         'width': 1.5,
    #         'dash_type': 'long_dash'
    #     }
    # })

    # lijnen toevoegen aan lijngrafiek
    count = 0
    for name, group in grouped:
        profielnaam = str(int(name))
        meetpunten = len(group['profielnummer'])

        # eerste profiel
        if count == 0:
            line_chart1.add_series({
                'name': 'profiel ' + profielnaam,

                'categories': '=Sheet1!B' + str(startpunt) + ':B' + str(meetpunten + 1),
                'values': '=Sheet1!C' + str(startpunt) + ':C' + str(meetpunten + 1),
                'line': {'width': 1}
            })
            count +=1
        # opvolgende profielen
        else:
            if count is not 0:
                line_chart1.add_series({
                    'name': 'profiel '+profielnaam,

                    'categories': '=Sheet1!B'+str(startpunt)+':B' + str(startpunt+meetpunten-1),
                    'values':     '=Sheet1!C'+str(startpunt)+':C' + str(startpunt+meetpunten-1),
                    'line': {'width': 1}
                })
        # startpunt verzetten
        startpunt += (meetpunten)



    ## toetshoogte toevoegen als horizontale lijn, deel 2 voor voorgrond-lijn
    # minimum = min(sorted['afstand'])
    # maximum = max(sorted['afstand'])
    # th = sorted[toetspeil].iloc[0]
    #
    # worksheet.write('K8', minimum)
    # worksheet.write('K9', maximum)
    # worksheet.write('K10', th)
    # worksheet.write('K11', th)
    #
    # line_chart1.add_series({
    #     'name': 'toetshoogte: ' + str(th) + ' m NAP',
    #
    #     'categories': '=Sheet1!$K$8:$K$9',
    #     'values': '=Sheet1!$K$10:$K$11',
    #     'line': {
    #         'color': 'red',
    #         'width': 1.5,
    #         'dash_type': 'long_dash'
    #     }
    # })

    # kolommen aanpassen
    line_chart1.set_title({'name': 'Overzicht profielen dijkvak '+id})
    line_chart1.set_x_axis({'name': 'Afstand [m]'})
    line_chart1.set_y_axis({'name': 'Hoogte [m NAP]'})
    line_chart1.set_x_axis({'interval_tick': 0.5})
    line_chart1.set_x_axis({'min': min_plot, 'max': max_plot})
    line_chart1.set_size({'width': 1000, 'height': 300})
    # line_chart1.set_style(1)
    worksheet.insert_chart('G3', line_chart1) # alleen toevoegen voor toetshoogte
    workbook.close()

    print '.xlsx-bestand gemaakt voor profielset'

def excel_writer_maatgevend(uitvoerpunten,code,excel,id,trajecten,toetspeil,min_plot,max_plot):

    # toetshoogte aan uitvoerpunten koppelen
    arcpy.JoinField_management(uitvoerpunten, code, trajecten, code, toetspeil)

    # binnenhalen van dataframe
    if toetspeil == 999:
        array = arcpy.da.FeatureClassToNumPyArray(uitvoerpunten,('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn', 'x', 'y'))
    else:
        array = arcpy.da.FeatureClassToNumPyArray(uitvoerpunten, ('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn', 'x', 'y', toetspeil))

    df = pd.DataFrame(array)
    df = df.dropna()
    sorted = df.sort_values(['profielnummer', 'afstand'], ascending=[True, True])

    # opbouw xlsx
    workbook = Workbook(excel)
    worksheet1 = workbook.add_worksheet('Overzicht')
    worksheet2 = workbook.add_worksheet()

    # stijl toevoegen voor headers
    bold = workbook.add_format({'bold': True})


    # schrijf kolomnamen
    worksheet2.write(0, 0, "Profielnummer", bold)
    worksheet2.write(0, 1, "Afstand [m]", bold)
    worksheet2.write(0, 2, "Hoogte AHN3 [m NAP]", bold)
    worksheet2.write(0, 3, "x [RD]", bold)
    worksheet2.write(0, 4, "y [RD]", bold)

    # schrijf kolommen vanuit df
    worksheet2.write_column('A2', sorted['profielnummer'])
    worksheet2.write_column('B2', sorted['afstand'])
    worksheet2.write_column('C2', sorted['z_ahn'])
    worksheet2.write_column('D2', sorted['x'])
    worksheet2.write_column('E2', sorted['y'])

    # groepeer per profielnummer
    grouped = sorted.groupby('profielnummer')

    # definieer startrij
    startpunt = 2


    # lege lijngrafiek invoegen met zowel afstand als hoogte als invoer
    line_chart1 = workbook.add_chart({'type': 'scatter',
                                 'subtype': 'straight'})

    ## toetshoogte, aan/uit
    # toetshoogte toevoegen als horizontale lijn, deel 1 voor legenda
    # minimum = min(sorted['afstand'])
    # maximum = max(sorted['afstand'])
    # th = sorted[toetspeil].iloc[0]
    #
    # worksheet.write('K8', minimum)
    # worksheet.write('K9', maximum)
    # worksheet.write('K10', th)
    # worksheet.write('K11', th)


    # line_chart1.add_series({
    #     'name': 'toetshoogte: ' + str(th) + ' m NAP',
    #
    #     'categories': '=Sheet1!$K$8:$K$9',
    #     'values': '=Sheet1!$K$10:$K$11',
    #     'line': {
    #         'color': 'red',
    #         'width': 1.5,
    #         'dash_type': 'long_dash'
    #     }
    # })

    # lijnen toevoegen aan lijngrafiek
    count = 0
    for name, group in grouped:
        profielnaam = str(int(name))
        meetpunten = len(group['profielnummer'])

        # eerste profiel
        if count == 0:
            line_chart1.add_series({
                'name': 'profiel ' + profielnaam,

                'categories': '=Sheet2!B' + str(startpunt) + ':B' + str(meetpunten + 1),
                'values': '=Sheet2!C' + str(startpunt) + ':C' + str(meetpunten + 1),
                'line': {'width': 1}
            })
            count +=1
        # opvolgende profielen
        else:
            if count is not 0 and name is not 9999:
                line_chart1.add_series({
                    'name': 'profiel '+profielnaam,

                    'categories': '=Sheet2!B'+str(startpunt)+':B' + str(startpunt+meetpunten-1),
                    'values':     '=Sheet2!C'+str(startpunt)+':C' + str(startpunt+meetpunten-1),
                    'line': {'width': 1}
                })
            if name == 9999:
                line_chart1.add_series({
                    'name': 'maatgevend profiel',

                    'categories': '=Sheet2!B'+str(startpunt)+':B' + str(startpunt+meetpunten-1),
                    'values':     '=Sheet2!C'+str(startpunt)+':C' + str(startpunt+meetpunten-1),
                    'line': {
                        'color': 'red',
                        'width': 3
                    }
                })
        # startpunt verzetten
        startpunt += (meetpunten)



    ## toetshoogte toevoegen als horizontale lijn, deel 2 voor voorgrond-lijn
    # minimum = min(sorted['afstand'])
    # maximum = max(sorted['afstand'])
    # th = sorted[toetspeil].iloc[0]
    #
    # worksheet.write('K8', minimum)
    # worksheet.write('K9', maximum)
    # worksheet.write('K10', th)
    # worksheet.write('K11', th)
    #
    # line_chart1.add_series({
    #     'name': 'toetshoogte: ' + str(th) + ' m NAP',
    #
    #     'categories': '=Sheet1!$K$8:$K$9',
    #     'values': '=Sheet1!$K$10:$K$11',
    #     'line': {
    #         'color': 'red',
    #         'width': 1.5,
    #         'dash_type': 'long_dash'
    #     }
    # })

    # kolommen aanpassen
    line_chart1.set_title({'name': 'Overzicht profielen prio-vak '+id})
    line_chart1.set_x_axis({'name': 'Afstand [m]'})
    line_chart1.set_y_axis({'name': 'Hoogte [m NAP]'})
    line_chart1.set_x_axis({'interval_tick': 0.5})
    line_chart1.set_x_axis({'min': min_plot, 'max': max_plot})
    line_chart1.set_size({'width': 1000, 'height': 300})
    # line_chart1.set_style(1)
    worksheet1.insert_chart('G3', line_chart1) # alleen toevoegen voor toetshoogte
    # worksheet2.insert_chart('G3', line_chart1) # test
    # worksheet2.hide()
    workbook.close()

    print '.xlsx-bestand gemaakt voor profielset'

def excel_writer_factsheets_main(uitvoerpunten,code,excel,id,trajecten,toetspeil,min_plot,max_plot,trajectlijn,img,percelen_zone):

    # toetshoogte aan uitvoerpunten koppelen
    arcpy.JoinField_management(uitvoerpunten, code, trajecten, code, toetspeil)

    # binnenhalen van dataframe
    if toetspeil == 999:
        array = arcpy.da.FeatureClassToNumPyArray(uitvoerpunten,('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn', 'x', 'y'))
    else:
        array = arcpy.da.FeatureClassToNumPyArray(uitvoerpunten, ('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn', 'x', 'y', toetspeil))

    df = pd.DataFrame(array)
    df = df.dropna()
    sorted = df.sort_values(['profielnummer', 'afstand'], ascending=[True, True])

    # opbouw xlsx
    workbook = Workbook(excel)
    worksheet1 = workbook.add_worksheet('Overzicht')
    worksheet2 = workbook.add_worksheet()
    worksheet3 = workbook.add_worksheet('Perceelgegevens')
    # stijl toevoegen voor headers
    bold = workbook.add_format({'bold': True})


    # schrijf kolomnamen
    worksheet2.write(0, 0, "Profielnummer", bold)
    worksheet2.write(0, 1, "Afstand [m]", bold)
    worksheet2.write(0, 2, "Hoogte AHN3 [m NAP]", bold)
    worksheet2.write(0, 3, "x [RD]", bold)
    worksheet2.write(0, 4, "y [RD]", bold)

    # schrijf kolommen vanuit df
    worksheet2.write_column('A2', sorted['profielnummer'])
    worksheet2.write_column('B2', sorted['afstand'])
    worksheet2.write_column('C2', sorted['z_ahn'])
    worksheet2.write_column('D2', sorted['x'])
    worksheet2.write_column('E2', sorted['y'])

    # groepeer per profielnummer
    grouped = sorted.groupby('profielnummer')

    # definieer startrij
    startpunt = 2


    # lege lijngrafiek invoegen met zowel afstand als hoogte als invoer
    line_chart1 = workbook.add_chart({'type': 'scatter',
                                 'subtype': 'straight'})

    ## toetshoogte, aan/uit
    # toetshoogte toevoegen als horizontale lijn, deel 1 voor legenda
    # minimum = min(sorted['afstand'])
    # maximum = max(sorted['afstand'])
    # th = sorted[toetspeil].iloc[0]
    #
    # worksheet.write('K8', minimum)
    # worksheet.write('K9', maximum)
    # worksheet.write('K10', th)
    # worksheet.write('K11', th)


    # line_chart1.add_series({
    #     'name': 'toetshoogte: ' + str(th) + ' m NAP',
    #
    #     'categories': '=Sheet1!$K$8:$K$9',
    #     'values': '=Sheet1!$K$10:$K$11',
    #     'line': {
    #         'color': 'red',
    #         'width': 1.5,
    #         'dash_type': 'long_dash'
    #     }
    # })

    # lijnen toevoegen aan lijngrafiek
    count = 0
    for name, group in grouped:
        profielnaam = str(int(name))
        meetpunten = len(group['profielnummer'])

        # eerste profiel
        if count == 0:
            line_chart1.add_series({
                'name': 'profiel ' + profielnaam,

                'categories': '=Sheet2!B' + str(startpunt) + ':B' + str(meetpunten + 1),
                'values': '=Sheet2!C' + str(startpunt) + ':C' + str(meetpunten + 1),
                'line': {'width': 1}
            })
            count +=1
        # opvolgende profielen
        else:
            if count is not 0 and name is not 9999:
                line_chart1.add_series({
                    'name': 'profiel '+profielnaam,

                    'categories': '=Sheet2!B'+str(startpunt)+':B' + str(startpunt+meetpunten-1),
                    'values':     '=Sheet2!C'+str(startpunt)+':C' + str(startpunt+meetpunten-1),
                    'line': {'width': 1}
                })
            if name == 9999:
                line_chart1.add_series({
                    'name': 'maatgevend profiel',

                    'categories': '=Sheet2!B'+str(startpunt)+':B' + str(startpunt+meetpunten-1),
                    'values':     '=Sheet2!C'+str(startpunt)+':C' + str(startpunt+meetpunten-1),
                    'line': {
                        'color': 'red',
                        'width': 3
                    }
                })
        # startpunt verzetten
        startpunt += (meetpunten)



    ## toetshoogte toevoegen als horizontale lijn, deel 2 voor voorgrond-lijn
    # minimum = min(sorted['afstand'])
    # maximum = max(sorted['afstand'])
    # th = sorted[toetspeil].iloc[0]
    #
    # worksheet.write('K8', minimum)
    # worksheet.write('K9', maximum)
    # worksheet.write('K10', th)
    # worksheet.write('K11', th)
    #
    # line_chart1.add_series({
    #     'name': 'toetshoogte: ' + str(th) + ' m NAP',
    #
    #     'categories': '=Sheet1!$K$8:$K$9',
    #     'values': '=Sheet1!$K$10:$K$11',
    #     'line': {
    #         'color': 'red',
    #         'width': 1.5,
    #         'dash_type': 'long_dash'
    #     }
    # })

    # kolommen aanpassen
    line_chart1.set_title({'name': 'Overzicht profielen prio-vak '+id})
    line_chart1.set_x_axis({'name': 'Afstand [m]'})
    line_chart1.set_y_axis({'name': 'Hoogte [m NAP]'})
    line_chart1.set_x_axis({'interval_tick': 0.5})
    line_chart1.set_x_axis({'min': min_plot, 'max': max_plot})
    line_chart1.set_size({'width': 1000, 'height': 300})
    # line_chart1.set_style(1)
    worksheet1.insert_chart('D24', line_chart1) # alleen toevoegen voor toetshoogte

    # worksheet2.insert_chart('G3', line_chart1) # test
    worksheet2.hide()

    # schrijf parameters uit trajectlijn naar worksheet1
    # format worksheet
    cell_format_title = workbook.add_format()
    cell_format_title.set_font_size(16)
    cell_format_title.set_bold()

    cell_format_sub = workbook.add_format()
    cell_format_sub.set_pattern(1)
    cell_format_sub.set_bg_color('#e6e65c')
    cell_format_sub.set_bold()

    # stel kolom breedtes in
    worksheet1.set_column(0, 0, 30)
    worksheet1.set_column(1, 1, 60)
    worksheet3.set_column(0, 6, 15)



    # search cursor om er doorheen te gaan en parameters eruit te halen

    worksheet1.write('A1', "Factsheet prio-vak "+str(id), cell_format_title)

    worksheet1.write('A3', "", cell_format_sub)
    worksheet1.write('B4', "", cell_format_sub)
    worksheet1.write('B11', "", cell_format_sub)
    worksheet1.write('B15', "", cell_format_sub)
    worksheet1.write('B20', "", cell_format_sub)
    worksheet1.write('B24', "", cell_format_sub)

    worksheet1.write('B3', "waarde",cell_format_sub)


    worksheet1.write('A4', "Algemeen",cell_format_sub)
    worksheet1.write('A5', "Vaknummer")
    worksheet1.write('A6', "Van dijkpaal")
    worksheet1.write('A7', "Tot dijkpaal")
    worksheet1.write('A8', "Vaklengte [m]")
    worksheet1.write('A9', "Laatste versterking [traject]")
    worksheet1.write('A10', "Laatste versterking [jaar]")

    worksheet1.write('A11', "Basisgegevens techniek",cell_format_sub)
    worksheet1.write('A12', "Dikte deklaag gemiddeld [m]")
    worksheet1.write('A13', "Dik deklaag variatie [m]")
    worksheet1.write('A14', "Deformatie gemiddeld[mm/jaar]")

    worksheet1.write('A15', "Basisgegevens conditionering",cell_format_sub)
    worksheet1.write('A16', "Huizen binnen teenlijn [aantal]")
    worksheet1.write('A17', "Huizen +20m teenlijn [aantal]")
    worksheet1.write('A18', "Percelen binnen zone.. [aantal]")
    worksheet1.write('A19', "Leidingen [m]")
    worksheet1.write('A20', "Natura 2000")

    worksheet1.write('A21', "Beoordeling techniek",cell_format_sub)
    worksheet1.write('A22', "STPH [beta]")
    worksheet1.write('A23', "STBI [beta]")
    worksheet1.write('A24', "GEKB [beta]")

    worksheet1.write('A25', "Ontwerpproces",cell_format_sub)
    worksheet1.write('A26', "Groep VVK")
    worksheet1.write('A27', "Maatregel VVK [soort]")
    worksheet1.write('A28', "Kosten VVK [*miljoen euro]")
    worksheet1.write('A29', "Extra sonderingen [aantal]")
    worksheet1.write('A30', "Extra boringen [aantal]")

    worksheet1.write('A31', "Geometrie")

    # maak array-pandas df van trajectlijn
    velden = ["prio_nummer","Van","Tot","Shape_Length","TRAJECT","OPLEVERING","gem_dpip","var_dpip","gem_zet","panden_dijkzone", "panden_dijkzone_bit",
              "lengte_kl", "extra_bo", "extra_so", "gekb_2023","stbi_2023","stph_2023","na2000","extra_inmeten","maatregel","kosten","groep","percelen_zone"]
    array_fact = arcpy.da.FeatureClassToNumPyArray(trajectlijn,velden)
    df_fact= pd.DataFrame(array_fact)


    nummer = df_fact['prio_nummer'].iloc[0]
    van = df_fact['Van'].iloc[0]
    tot = df_fact['Tot'].iloc[0]
    lengte = int(df_fact['Shape_Length'].iloc[0])

    # test op NaN

    traject = df_fact['TRAJECT'].iloc[0]
    oplevering = df_fact['OPLEVERING'].iloc[0]
    gdpip = df_fact['gem_dpip'].iloc[0]
    vdpip = df_fact['var_dpip'].iloc[0]
    gzet = df_fact['gem_zet'].iloc[0]
    pdijk = df_fact['panden_dijkzone'].iloc[0]
    pbit = df_fact['panden_dijkzone_bit'].iloc[0]
    lengtekl = df_fact['lengte_kl'].iloc[0]
    na2000 = df_fact['na2000'].iloc[0]
    stph = df_fact['stph_2023'].iloc[0]
    stbi = df_fact['stbi_2023'].iloc[0]
    gekb = df_fact['gekb_2023'].iloc[0]
    groep = df_fact['groep'].iloc[0]
    kosten = df_fact['kosten'].iloc[0]
    maatregel = df_fact['maatregel'].iloc[0]
    extrabo = int(df_fact['extra_bo'].iloc[0])
    extraso = int(df_fact['extra_so'].iloc[0])
    extrameet= df_fact['extra_inmeten'].iloc[0]
    percelen = df_fact['percelen_zone'].iloc[0]




    # schrijf parameters naar de excelcellen
    worksheet1.write('B5',  nummer)
    worksheet1.write('B6', van)
    worksheet1.write('B7', tot)
    worksheet1.write('B8', str(lengte))

    if str(traject) == "None":
        worksheet1.write('B9', "n.v.t.")
    else:
        worksheet1.write('B9', traject)


    if str(oplevering) == "None":
        worksheet1.write('B10', "n.v.t.")
    else:
        worksheet1.write('B10', oplevering)

    if pd.isna(gdpip) == True:
        worksheet1.write('B12', "n.v.t.")
    else:
        gdpip = round(gdpip, 1)
        worksheet1.write('B12', str(gdpip))

    if pd.isna(vdpip) == True:
        worksheet1.write('B13', "n.v.t.")
    else:
        vdpip = round(vdpip, 1)
        worksheet1.write('B13', str(vdpip))


    if pd.isna(gzet) == True:
        worksheet1.write('B14', "n.v.t.")
    else:
        gzet = round(gzet, 1)
        worksheet1.write('B14', str(gzet))




    if pdijk > 0:
        pdijk = int(pdijk)
        worksheet1.write('B16', str(pdijk))
    else:
        worksheet1.write('B16', "n.v.t.")

    if pbit > 0:
        pbit = int(pbit)
        worksheet1.write('B17', str(pbit))
    else:
        worksheet1.write('B17', "n.v.t.")

    if percelen > 0:
        percelen = int(percelen)
        worksheet1.write('B18', str(percelen))
    else:
        worksheet1.write('B18', "n.v.t.")


    if pd.isna(lengtekl) == True:
        worksheet1.write('B19', "n.v.t.")
    else:
        lengtekl = int(lengtekl)
        worksheet1.write('B19', str(lengtekl))


    if na2000 == "Ja":
        worksheet1.write('B20', "Aanwezig")
    else:
        worksheet1.write('B20', "n.v.t.")

    if pd.isna(stph) == True:
        worksheet1.write('B22', "n.v.t.")
    else:
        stph = round(stph, 1)
        worksheet1.write('B22', str(stph))

    if pd.isna(stbi) == True:
        worksheet1.write('B23', "n.v.t.")
    else:
        stbi = round(stbi, 1)
        worksheet1.write('B23', str(stbi))

    if pd.isna(gekb) == True:
        worksheet1.write('B24', "n.v.t.")
    else:
        gekb = round(gekb, 1)
        worksheet1.write('B24', str(gekb))



    if pd.isna(groep) == True:
        worksheet1.write('B26', "n.v.t.")
    else:
        worksheet1.write('B26', str(groep))

    if pd.isna(maatregel) == True:
        worksheet1.write('B27', "n.v.t.")
    else:
        worksheet1.write('B27', str(maatregel))



    if pd.isna(kosten) == True:
        worksheet1.write('B28', "Onbekend")
    else:
        worksheet1.write('B28', str(kosten))



    if pd.isna(extraso) == True or extraso < 1:
        worksheet1.write('B29', "n.v.t.")
    else:
        extrago = int(extraso)
        worksheet1.write('B29', str(extraso))

    if pd.isna(extrabo) == True or extrabo < 1:
        worksheet1.write('B30', "n.v.t.")
    else:
        extrago = int(extrabo)
        worksheet1.write('B30', str(extrabo))

    if extrameet == "Ja":
        worksheet1.write('B31', "Extra inmetingen vereist")
    else:
        worksheet1.write('B31', "Geen inmetingen vereist")


    # insert plot vanuit arcmap
    worksheet1.insert_image('D3', img)




    ## voeg perceeldata toe aan nieuwe sheet

    # data binnenhalen
    velden_percelen = ["OBJECTID","Huisnummer","Huisletter","Postcode","OpenbareRuimteNaam","WoonplaatsNaam"]
    array_percelen = arcpy.da.FeatureClassToNumPyArray(percelen_zone, velden_percelen, null_value=-9999)
    df = pd.DataFrame(array_percelen)
    df_percelen = df.sort_values('OpenbareRuimteNaam', ascending=False)

    # kolomnamen
    worksheet3.write('A1', "OBJECTID_gis", cell_format_sub)
    worksheet3.write('B1', "Straatnaam", cell_format_sub)
    worksheet3.write('C1', "Huisnummer", cell_format_sub)
    worksheet3.write('D1', "Huisletter", cell_format_sub)
    worksheet3.write('E1', "Postcode", cell_format_sub)
    worksheet3.write('F1', "Plaatsnaam", cell_format_sub)

    # schrijven kolommen
    worksheet3.write_column('A2', df_percelen['OBJECTID'])
    worksheet3.write_column('B2', df_percelen['OpenbareRuimteNaam'])
    worksheet3.write_column('C2', df_percelen['Huisnummer'])
    worksheet3.write_column('D2', df_percelen['Huisletter'])
    worksheet3.write_column('E2', df_percelen['Postcode'])
    worksheet3.write_column('F2', df_percelen['WoonplaatsNaam'])


    workbook.close()
    del df
    del df_percelen
    del df_fact




    print '.xlsx-bestand gemaakt voor factsheet'




def binnenteenbepalen(invoer, code, min_achterland, max_achterland, uitvoer_binnenteen, min_afstand,
                      max_afstand,uitvoer_binnenkruin):

    if arcpy.Exists(uitvoer_binnenkruin):

        array = arcpy.da.FeatureClassToNumPyArray(invoer, ('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn'))
        df = pd.DataFrame(array)
        df2 = df.dropna()
        sorted = df2.sort_values(['profielnummer', 'afstand'], ascending=[True, True])
        grouped = sorted.groupby('profielnummer')

        # lijst OBJECTIDs
        list_id_bit = []

        # binnenhalen binnenkruindata
        array_bik = arcpy.da.FeatureClassToNumPyArray(uitvoer_binnenkruin, ('OBJECTID', 'profielnummer', code, 'afstand', 'z_ahn'))
        df_bik = pd.DataFrame(array_bik)

        # print df_bik
        # print(df_bik.loc[df_bik['afstand'] == 4.5])

        for name, group in grouped:

            bik_afstand = df_bik.loc[df_bik['profielnummer'] == name, 'afstand']
            for item in bik_afstand.values:
                bik = item

            # maximale kruinhoogte
            max_kruin = max(group['z_ahn'])
            landzijde = group.sort_values(['afstand'], ascending=False)  # afnemend, landzijde
            rivierzijde = group.sort_values(['afstand'], ascending=True)  # toenemend, rivierzijde

            # maaiveldhoogte dijk
            mv_dijk_lijst = []
            for index, row in landzijde.iterrows():
                if row['afstand'] > min_afstand and row['afstand'] < max_afstand:
                    mv_dijk_lijst.append(row['z_ahn'])

            if mv_dijk_lijst:
                mv_dijk = average(mv_dijk_lijst)
                # print mv_dijk

            # maaiveldhoogte achterland
            mv_achterland_lijst = []
            for index, row in landzijde.iterrows():
                if row['afstand'] > min_achterland and row['afstand'] < max_achterland:
                    mv_achterland_lijst.append(row['z_ahn'])

            if mv_achterland_lijst:
                mv_achterland = average(mv_achterland_lijst)
                # print mv_achterland

            x1 = group['afstand']
            y1 = group['z_ahn']

            # extra, toevoegen indien nodig
            # f = interp1d(x1, y1, kind='linear')
            # x_new = np.linspace(min(x1), max(x1), 200)

            # plot, standaard uit
            # fig = plt.figure(figsize=(25, 2))
            # ax = fig.add_subplot(111)
            # ax.plot(x1, y1, linewidth=2, color="red")
            # ax.plot(x_new, f(x_new), linewidth=2, color="blue")
            # ax.plot(x_new, f(x_new), 'o', color="blue")
            # ax.plot(x1, y1, 'o', color="red")

            # aanpassen van dataframe voor bepaling binnenteen
            df = rivierzijde
            df['next_afstand'] = df['afstand'].shift(-1)
            df['next_z'] = df['z_ahn'].shift(-1)
            df['talud'] = abs((df['z_ahn'] - df['next_z']) / (df['afstand'] - df['next_afstand']))
            df['max_talud'] = df.iloc[:, 7].rolling(window=3).mean()
            df['next_max_talud'] = df['max_talud'].shift(-1)


            # for index, row in df.iterrows():
            #     print df_bik.loc['afstand']
                # if df_bik.loc[df_bik['afstand'] == row['afstand']]:
                #     print row['afstand']
                # afstand = row['afstand']
                # if afstand > min_afstand and afstand < max_afstand:
                #     df_profiel_tester.loc[row['OBJECTID']] = name, row['middelpunt_max']

            try:
                mv_dijk, mv_achterland
                for index, row in df.iterrows():
                    if mv_dijk > mv_achterland and row['afstand'] - row['next_afstand'] <= 2 and row['afstand'] > 3 and \
                            row['next_max_talud'] < row['max_talud'] and row['afstand'] > bik:
                        list_id_bit.append(row['OBJECTID'])
                        # ax.plot(row['next_afstand'], row['next_z'], 'o', markersize=12)
                        break
            except NameError:
                list_id_bit = []


            # extra, knikpunten o.b.v. RDP
            # tolerance = 0.1
            # min_angle = np.pi * 0.02
            # points = np.c_[x1, y1]
            # simplified_trajectory = np.array(rdp.rdp(points.tolist(), tolerance))
            # sx, sy = simplified_trajectory.T

            # vectoren berekenen
            # directions = np.diff(simplified_trajectory, axis=0)
            # theta = angle(directions)

            # selecteer index van punten met grootste theta, hoge theta > maximale verandering in richting
            # idx = np.where(theta > min_angle)[0] + 1

            # plot knikpunten met versimpeld profiel

            # ax.plot(sx, sy, 'gx-', label='simplified trajectory')
            # ax.plot(sx[idx], sy[idx], 'ro', markersize=10, label='turning points'

            # plt.show()

        # wegschrijven naar gis
        if list_id_bit:
            arcpy.MakeFeatureLayer_management(invoer, 'punten_binnenteen_temp')
            punten_bit = arcpy.SelectLayerByAttribute_management('punten_binnenteen_temp', "ADD_TO_SELECTION",
                                                                 "OBJECTID in (" + str(list_id_bit)[1:-1] + ")")
            arcpy.CopyFeatures_management('punten_binnenteen_temp', uitvoer_binnenteen)
        print 'Binnenteen bepaald'
    else:
        pass
def koppeling_hbn_hdsr(profielen,toetspeil):


    bestaande_velden = arcpy.ListFields(profielen)
    te_verwijderen = ['kruin2014_th']
    for field in bestaande_velden:
        if field.name in te_verwijderen:
            arcpy.DeleteField_management(profielen, field.name)

    arcpy.AddField_management(profielen, "kruin2014_th", "DOUBLE", 2, field_is_nullable="NULLABLE")



    with arcpy.da.UpdateCursor(profielen, ['maxKruinhoogte2014', toetspeil, 'kruin2014_th']) as cursor:

        for row in cursor:
            if row[0] is not None and row[0] >= row[1]:
                verschil = abs(row[0] - row[1])
                row[2] = round(verschil,2)
                cursor.updateRow(row)
            else:
                if row[0] is not None and row[0]<row[1]:
                    verschil = abs(row[0] - row[1])
                    row[2] = -round(verschil,2)
                    cursor.updateRow(row)

    print "Kruinhoogte-hbn berekend"


def profielen_op_lijn(profiel_interval,profiel_lengte_land,profiel_lengte_rivier,trajectlijn,code, profielen):
    # traject to points
    arcpy.GeneratePointsAlongLines_management(trajectlijn, 'traject_punten', 'DISTANCE', Distance=profiel_interval, Include_End_Points='END_POINTS')
    arcpy.AddField_management('traject_punten', "profielnummer", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management('traject_punten', "lengte_landzijde", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management('traject_punten', "lengte_rivierzijde", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management('traject_punten', "profielnummer", '!OBJECTID!', "PYTHON")
    arcpy.CalculateField_management('traject_punten', "lengte_landzijde", profiel_lengte_land, "PYTHON")
    arcpy.CalculateField_management('traject_punten', "lengte_rivierzijde", profiel_lengte_rivier, "PYTHON")

    # route voor trajectlijn
    # arcpy.CreateRoutes_lr(trajectlijn, code, "route_traject", "LENGTH", "", "", "UPPER_LEFT", "1", "0", "IGNORE", "INDEX")

    existing_fields = arcpy.ListFields(trajectlijn)
    needed_fields = ['OBJECTID', 'SHAPE', 'SHAPE_Length','Shape','Shape_Length',code]
    for field in existing_fields:
        if field.name not in needed_fields:
            arcpy.DeleteField_management(trajectlijn, field.name)
    arcpy.AddField_management(trajectlijn, "van", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management(trajectlijn, "tot", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management(trajectlijn, "van", 0, "PYTHON")
    arcpy.CalculateField_management(trajectlijn, "tot", "!Shape_Length!", "PYTHON")
    arcpy.CreateRoutes_lr(trajectlijn, code, 'route_traject', "TWO_FIELDS", "van", "tot", "", "1",
                          "0", "IGNORE", "INDEX")


    # locate profielpunten
    arcpy.LocateFeaturesAlongRoutes_lr('traject_punten', 'route_traject', code, "1.5 Meters", 'tabel_traject_punten',
                                       "RID POINT MEAS", "FIRST", "DISTANCE", "ZERO", "FIELDS",
                                       "M_DIRECTON")

    # offset rivierdeel profiel
    arcpy.MakeRouteEventLayer_lr('route_traject', code, 'tabel_traject_punten', "rid POINT meas", 'deel_rivier',
                                 "lengte_rivierzijde", "NO_ERROR_FIELD", "NO_ANGLE_FIELD", "NORMAL", "ANGLE", "RIGHT",
                                 "POINT")

    arcpy.MakeRouteEventLayer_lr('route_traject', code, 'tabel_traject_punten', "rid POINT meas", 'deel_land',
                                 "lengte_landzijde", "NO_ERROR_FIELD", "NO_ANGLE_FIELD", "NORMAL", "ANGLE", "LEFT",
                                 "POINT")
    # temp inzicht layer
    arcpy.CopyFeatures_management('deel_rivier', "temp_rivierdeel")
    arcpy.CopyFeatures_management('deel_land', "temp_landdeel")
    arcpy.AddField_management('temp_rivierdeel', "id", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management('temp_landdeel', "id", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management('temp_rivierdeel', "id", 2, "PYTHON")
    arcpy.CalculateField_management('temp_landdeel', "id", 1, "PYTHON")





    arcpy.Merge_management("'temp_rivierdeel';'temp_landdeel'", 'merge_profielpunten')
    arcpy.PointsToLine_management('merge_profielpunten', profielen, "profielnummer", "id", "NO_CLOSE")

    arcpy.SpatialJoin_analysis(profielen, trajectlijn, 'profielen_temp', "JOIN_ONE_TO_ONE", "KEEP_ALL", match_option="INTERSECT")
    arcpy.CopyFeatures_management('profielen_temp', profielen)

    # arcpy.AddField_management(profielen, "van", "DOUBLE", 2, field_is_nullable="NULLABLE")
    # arcpy.AddField_management(trajectlijn, "tot", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management(profielen, "van", 0, "PYTHON")
    arcpy.CalculateField_management(profielen, "tot", "!Shape_Length!", "PYTHON")

    # arcpy.FlipLine_edit(profielen)

    print 'profielen gemaakt op trajectlijn'

def generate_profiles_onpoints(traject_punten,trajectlijn,profielen,code):
    # traject to points
    traject_punten = traject_punten
    trajectlijn = trajectlijn
    profiel_lengte_rivier = 100
    profiel_lengte_land = 100
    profielen = profielen
    code = code

    existing_fields = arcpy.ListFields(profielen)
    needed_fields = ["profielnummer","lengte_landzijde","lengte_rivierzijde"]
    for field in existing_fields:
        if field.name in needed_fields:
            arcpy.DeleteField_management(trajectlijn, field.name)
    arcpy.AddField_management(traject_punten, "profielnummer", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management(traject_punten, "lengte_landzijde", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management(traject_punten, "lengte_rivierzijde", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management(traject_punten, "profielnummer", '!OBJECTID!', "PYTHON")
    arcpy.CalculateField_management(traject_punten, "lengte_landzijde", profiel_lengte_land, "PYTHON")
    arcpy.CalculateField_management(traject_punten, "lengte_rivierzijde", profiel_lengte_rivier, "PYTHON")

    # route voor trajectlijn
    # arcpy.CreateRoutes_lr(trajectlijn, code, "route_traject", "LENGTH", "", "", "UPPER_LEFT", "1", "0", "IGNORE", "INDEX")

    existing_fields = arcpy.ListFields(trajectlijn)
    needed_fields = ['OBJECTID', 'SHAPE', 'SHAPE_Length','Shape','Shape_Length',code]
    for field in existing_fields:
        if field.name not in needed_fields:
            arcpy.DeleteField_management(trajectlijn, field.name)
    arcpy.AddField_management(trajectlijn, "van", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management(trajectlijn, "tot", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management(trajectlijn, "van", 0, "PYTHON")
    arcpy.CalculateField_management(trajectlijn, "tot", "!Shape_Length!", "PYTHON")
    arcpy.CreateRoutes_lr(trajectlijn, code, 'route_traject', "TWO_FIELDS", "van", "tot", "", "1",
                          "0", "IGNORE", "INDEX")


    # locate profielpunten
    arcpy.LocateFeaturesAlongRoutes_lr(traject_punten, 'route_traject', code, "1.5 Meters", 'tabel_traject_punten',
                                       "RID POINT MEAS", "FIRST", "DISTANCE", "ZERO", "FIELDS",
                                       "M_DIRECTON")

    # offset rivierdeel profiel
    arcpy.MakeRouteEventLayer_lr('route_traject', code, 'tabel_traject_punten', "rid POINT meas", 'deel_rivier',
                                 "lengte_rivierzijde", "NO_ERROR_FIELD", "NO_ANGLE_FIELD", "NORMAL", "ANGLE", "RIGHT",
                                 "POINT")

    arcpy.MakeRouteEventLayer_lr('route_traject', code, 'tabel_traject_punten', "rid POINT meas", 'deel_land',
                                 "lengte_landzijde", "NO_ERROR_FIELD", "NO_ANGLE_FIELD", "NORMAL", "ANGLE", "LEFT",
                                 "POINT")
    # temp inzicht layer
    arcpy.CopyFeatures_management('deel_rivier', "temp_rivierdeel")
    arcpy.CopyFeatures_management('deel_land', "temp_landdeel")
    arcpy.AddField_management('temp_rivierdeel', "id", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management('temp_landdeel', "id", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.CalculateField_management('temp_rivierdeel', "id", 2, "PYTHON")
    arcpy.CalculateField_management('temp_landdeel', "id", 1, "PYTHON")





    arcpy.Merge_management("'temp_rivierdeel';'temp_landdeel'", 'merge_profielpunten')
    arcpy.PointsToLine_management('merge_profielpunten', profielen, "profielnummer", "id", "NO_CLOSE")

    # arcpy.CalculateField_management(profielen, "van", 0, "PYTHON")
    # arcpy.CalculateField_management(profielen, "tot", "!Shape_Length!", "PYTHON")

    # arcpy.SpatialJoin_analysis(profielen, trajectlijn, 'profielen_temp', "JOIN_ONE_TO_ONE", "KEEP_ALL", match_option="INTERSECT")
    # arcpy.CopyFeatures_management('profielen_temp', profielen)
    # arcpy.FlipLine_edit(profielen)

    print 'profielen gemaakt op trajectlijn'


def bereken_restlevensduur(profielen,bodemdalingskaart,afstand_zichtjaar,toetspeil):
    # profiellijnen middenpunt
    arcpy.FeatureToPoint_management(profielen, 'profielen_temp','CENTROID')

    # middenpunt z-waarde bodemdaling
    arcpy.CheckOutExtension("Spatial")
    ExtractValuesToPoints('profielen_temp', bodemdalingskaart, 'profielen_temp_z',
                          "NONE", "VALUE_ONLY")

    # Pas het veld 'RASTERVALU' aan naar 'z_ahn'
    bestaande_velden = arcpy.ListFields('profielen_temp_z')
    te_verwijderen = ['bdMmj', 'rldJaar']
    for field in bestaande_velden:
        if field.name in te_verwijderen:
            arcpy.DeleteField_management('profielen_temp_z', field.name)
    arcpy.AlterField_management('profielen_temp_z', 'RASTERVALU', 'bdMmj')
    del bestaande_velden

    # punt toevoegen aan profielen
    bestaande_velden = arcpy.ListFields(profielen)
    te_verwijderen = ['bdMmj', 'rldJaar']
    for field in bestaande_velden:
        if field.name in te_verwijderen:
            arcpy.DeleteField_management(profielen, field.name)
    arcpy.JoinField_management(profielen, 'profielnummer', 'profielen_temp_z', 'profielnummer', 'bdMmj')



    # maximale kruinhoogte aanpassen aan bodemdaling, 'bd_' voor velden
    bestaande_velden = arcpy.ListFields(profielen)
    te_verwijderen = ['maxKruinhoogte2024','kruin2024_th']
    for field in bestaande_velden:
        if field.name in te_verwijderen:
            arcpy.DeleteField_management(profielen, field.name)
    arcpy.AddField_management(profielen, "maxKruinhoogte2024", "DOUBLE", 2, field_is_nullable="NULLABLE")
    arcpy.AddField_management(profielen, "kruin2024_th", "DOUBLE", 2, field_is_nullable="NULLABLE")

    # bereken maximale kruinhoogte-bodemdaling per x jaren
    with arcpy.da.UpdateCursor(profielen, ['profielnummer','bdMmj','maxKruinhoogte2014', 'maxKruinhoogte2024']) as cursor:
        for row in cursor:
            if row[2] is not None and row[1] < 0:
                bd_meters = row[1]/1000
                row[3] = round(row[2]-abs(afstand_zichtjaar*bd_meters),2)
                cursor.updateRow(row)
            elif row[2] is not None and row[1] >= 0:
                row[3] = row[2]
                row[1] = 0
                cursor.updateRow(row)
            elif row[2] is None and row[1] >= 0:
                row[1] = 0
                cursor.updateRow(row)
            else:
                pass


    # bereken verschil maximale kruinhoogte-bodemdaling per x jaren met hbn
    with arcpy.da.UpdateCursor(profielen, [toetspeil,'bdMmj','maxKruinhoogte2024','kruin2024_th']) as cursor:
        for row in cursor:
            if row[2] is not None and row[2] >= row[0]:
                verschil = abs(row[2] - row[0])
                row[3] = round(verschil,2)
                cursor.updateRow(row)
            else:
                if row[2] is not None and row[2]<row[0]:
                    verschil = abs(row[2] - row[0])
                    row[3] = -round(verschil,2)
                    cursor.updateRow(row)


    # restlevensduur berekenen
    arcpy.AddField_management(profielen, "rldJaar", "DOUBLE")
    with arcpy.da.UpdateCursor(profielen, ('bdMmj','rldJaar','kruin2024_th')) as cursor:
        for row in cursor:
            if row[2] is not None and row[0] is not None:
                resthoogte = row[2]
                if resthoogte <= 0:
                    row[1] = 0
                elif row[0] < 0:
                    resthoogte_mm = row[2]*1000
                    bodemdaling_mm = abs(row[0])
                    row[1] = round(resthoogte_mm/bodemdaling_mm,1)
                else:
                    if row[0]==0:
                        row[1] =999
                cursor.updateRow(row)

    print 'Restlevensduur berekend'


def excelWriterTraject(uitvoerpunten,excel, veldnamen):
    # df van profielpunten
    array = arcpy.da.FeatureClassToNumPyArray(uitvoerpunten, veldnamen)
    df = pd.DataFrame(array)

    #export excel
    df.to_excel(excel)  

    print ("Excel gemaakt van profieldata")