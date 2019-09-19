import os
import xlrd
from openpyxl import load_workbook


workbook = xlrd.open_workbook("C:/Users/vince/Desktop/uitvoer_script/safe_resultaten_gekb_stph.xlsx")
worksheet = workbook.sheet_by_name("stph_overzicht")


hbn2015 = worksheet.col_values(15, 1)
#print hbn2015

for root, dirs, files in os.walk("C:/Users/vince/Desktop/uitvoer_script/2015"):
    xlsfiles=[ _ for _ in files if _.endswith('.xlsx') ]
    for xlsfile in xlsfiles:
        wb = load_workbook(os.path.join(root,xlsfile))
        ws= wb["Scenario_1"]
        ws['C26'] = 100
        wb.save("C:/Users/vince/Desktop/2015/uitvoer_script/2015_aangepast/"+xlsfile)



# c0 = [worksheet.row_values(i)[0] for i in range(worksheet.nrows) if worksheet.row_values(i)[0]]
#kv = worksheet.cell(33, 6)
#ov = worksheet.cell(34, 6)


#wb = load_workbook("C:/Users/vince/Desktop/2015/16-3_AW211..xlsx")
#ws = wb["Scenario_1"]
#ws['C26'] = 100
#wb.save("C:/Users/vince/Desktop/2015/16-3_AW211..xlsx")