import os
import xlrd
from openpyxl import load_workbook

#define results for input .xlsx
workbook = xlrd.open_workbook("C:/Users/vince/Desktop/uitvoer_script/safe_resultaten_gekb_stph.xlsx")
worksheet = workbook.sheet_by_name("stph_overzicht")

voegnaam = worksheet.col_values(1, 1)
hbn2015 = worksheet.col_values(21, 1)


d = dict(zip(voegnaam, hbn2015))

#for key, value in d.items():
    #if key == filename:
        #print "true"
    #else: print "not true"

#iterate over stph files in folder and replace value
for root, dirs, files in os.walk("C:/Users/vince/Desktop/uitvoer_script/test"):
    xlsfiles=[ _ for _ in files if _.endswith('.xlsx') ]
    for xlsfile in xlsfiles:
        wb = load_workbook(os.path.join(root,xlsfile))
        ws= wb["Scenario_1"]

        for key, value in d.items():
            if key == xlsfile:
                ws['C26'] = value

        wb.save("C:/Users/vince/Desktop/uitvoer_script/test/"+xlsfile)


#extract calculated values from separate .xlsx and combine results


