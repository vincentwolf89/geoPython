import gc
gc.collect()
import arcpy
import math
from arcpy.sa import *
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from itertools import groupby
from mpltools import color
from os import path, listdir

from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter




arcpy.env.overwriteOutput = True

# definieer .gdb waarin gewerkt wordt
arcpy.env.workspace = r"C:\Users\vince\Desktop\GIS\profielen_vergelijken.gdb"

# join profielen to dijkvakindeling


profielset = "profielen_safe_25m"
dijkvakindeling = r"C:\Users\vince\Desktop\GIS\data.gdb\dijkvakindeling_stbi_juni_2019"
bovenlijn = r"C:\Users\vince\Desktop\GIS\profielscript_final.gdb\bovenlijn"
dijkpalen = r"C:\Users\vince\Desktop\GIS\data.gdb\dijkpalen_traject"
uitvoer_profielset_1 = 'profielen_uitvoer_1'
uitvoer_profielset_2 = 'profielen_uitvoer_2'
excel = r"C:\Users\vince\Desktop\profielen_safe_25m_per_dv.xlsx"
def join_dv_indeling():

    invoer = profielset
    joins = dijkvakindeling # dijkvakindeling
    uitvoer = uitvoer_profielset_1

    # geef aan welke velden behouden blijven
    fieldmappings = arcpy.FieldMappings()

    # Add all fields from inputs.
    fieldmappings.addTable(invoer)
    fieldmappings.addTable(joins)

    velden = ['profiel_id', 'dijktraject','dv_nummer','koppel_id','locatie','dijkpaalnummer']

    # definieer te behouden velden
    keepers = velden

    # verwijder overige velden
    for field in fieldmappings.fields:
        if field.name not in keepers:
            fieldmappings.removeFieldMap(fieldmappings.findFieldMapIndex(field.name))

    # voer de spatial join uit
    arcpy.SpatialJoin_analysis(invoer, joins, uitvoer, "#", "#", fieldmappings)

    del invoer, joins, uitvoer, fieldmappings, keepers
    print "Koppeling met dijkvakindeling gemaakt"

def join_dijkpalen():
    invoer = uitvoer_profielset_1
    joins = dijkpalen # dijkpalen 100 m
    uitvoer = uitvoer_profielset_2

    # geef aan welke velden behouden blijven
    fieldmappings = arcpy.FieldMappings()

    # Add all fields from inputs.
    fieldmappings.addTable(invoer)
    fieldmappings.addTable(joins)

    velden = ['profiel_id', 'dijktraject','dv_nummer','koppel_id','dijkpaalnummer', 'locatie']

    # definieer te behouden velden
    keepers = velden

    # verwijder overige velden
    for field in fieldmappings.fields:
        if field.name not in keepers:
            fieldmappings.removeFieldMap(fieldmappings.findFieldMapIndex(field.name))

    # voer de spatial join uit
    arcpy.SpatialJoin_analysis(invoer, joins, uitvoer,"JOIN_ONE_TO_ONE", "KEEP_ALL", fieldmappings, "CLOSEST", "", "")



    del invoer, joins, uitvoer, fieldmappings, keepers
    print "Koppeling met dijkpalen gemaakt"

def generate_points():
    # Set local variables
    invoer = uitvoer_profielset_1
    uitvoer_punten = 'profielen_punten'


    # Execute GeneratePointsAlongLines by distance
    arcpy.GeneratePointsAlongLines_management(invoer, uitvoer_punten, 'DISTANCE',
                                            Distance= 2)
    print "Punten op de profielen gemaakt"
    del invoer, uitvoer_punten

def values_points():
    # Set local variables
    invoer = "profielen_punten"
    inRaster = "C:/Users/vince/Desktop/GIS/losse rasters/ahn3clip/ahn3clip_safe"
    uitvoer = "profielen_punten_z"

    # Check out the ArcGIS Spatial Analyst extension license
    arcpy.CheckOutExtension("Spatial")

    # Execute ExtractValuesToPoints
    ExtractValuesToPoints(invoer, inRaster, uitvoer,
                          "INTERPOLATE", "VALUE_ONLY")
    print "Z-waarde aan punten gekoppeld"
    del invoer, uitvoer

def velden():

    arcpy.env.outputCoordinateSystem = arcpy.Describe("profielen_punten_z").spatialReference
    # Set local variables
    in_features = "profielen_punten_z"
    properties = "POINT_X_Y_Z_M"
    length_unit = ""
    area_unit = ""
    coordinate_system = ""

    # Generate the extent coordinates using Add Geometry Properties tool
    arcpy.AddGeometryAttributes_management(in_features, properties, length_unit,
                                           area_unit,
                                           coordinate_system)


def koppel_centerline():
    invoer = "profielen_punten_z"
    joins = dijkvakindeling
    uitvoer = "profielen_punten_z_join1"


    # pas de veldnamen aan

    hoogte = 'RASTERVALU'
    hoogte_nieuw = 'z_ahn'

    x_oud = 'POINT_X'
    x_nieuw = 'x'

    y_oud = 'POINT_Y'
    y_nieuw = 'y'

    arcpy.AlterField_management(invoer, hoogte, hoogte_nieuw)
    arcpy.AlterField_management(invoer, x_oud, x_nieuw)
    arcpy.AlterField_management(invoer, y_oud, y_nieuw)


    # spatial join punten 2
    velden = ['z_ahn', 'type_lijn', 'x', 'y', 'profiel_id', 'dijktraject', 'dv_nummer', 'koppel_id', 'dijkpaalnummer', 'locatie']  # defineeer de te behouden velden
    fieldmappings = arcpy.FieldMappings()

    # stel de fieldmapping in
    fieldmappings.addTable(invoer)
    fieldmappings.addTable(joins)
    keepers = velden

    # verwijder de niet-benodigde velden
    for field in fieldmappings.fields:
        if field.name not in keepers:
            fieldmappings.removeFieldMap(fieldmappings.findFieldMapIndex(field.name))
    # voer de spatial join uit
    arcpy.SpatialJoin_analysis(invoer, joins, uitvoer, "#", "#", fieldmappings, match_option="INTERSECT",
                               search_radius=1)
    print "Koppeling met centerline(dijkvakindeling) gemaakt"

def knip_profielen():
  #intersect with dv indeling --> point
  inFeatures = [profielset, dijkvakindeling]
  intersectOutput = "nul_punten_profielen"
  clusterTolerance = 0.5
  arcpy.Intersect_analysis(inFeatures, intersectOutput, "", clusterTolerance, "point")

  #split profielen voor koppeling
  inFeatures = profielset
  pointFeatures = "nul_punten_profielen"
  outFeatureclass = "profielknips"
  searchRadius = "1"
  arcpy.SplitLineAtPoint_management(inFeatures, pointFeatures, outFeatureclass, searchRadius)

  #koppel aan profielknips
  targetFeatures = "profielknips"
  joinFeatures = bovenlijn
  outfc = "profielen_bovenlijn"

  velden = ['profiel_id','rivierzijde']  # defineeer de te behouden velden
  fieldmappings = arcpy.FieldMappings()

  fieldmappings.addTable(targetFeatures)
  fieldmappings.addTable(joinFeatures)
  keepers = velden

  for field in fieldmappings.fields:
    if field.name not in keepers:
      fieldmappings.removeFieldMap(fieldmappings.findFieldMapIndex(field.name))

  arcpy.SpatialJoin_analysis(targetFeatures, joinFeatures, outfc, "#", "#", fieldmappings, match_option="INTERSECT",
                             search_radius=1)

  print "Profielen gesplit voor goede koppeling."


def koppel_bovenlijn():
    targetFeatures = "profielen_punten_z_join1"
    joinFeatures = "profielen_bovenlijn"
    outfc = "profielen_punten_z_join2"


    # spatial join punten 2
    velden = ['z_ahn', 'type_lijn', 'x', 'y', 'profiel_id', 'rivierzijde','dijktraject', 'dv_nummer', 'koppel_id', 'dijkpaalnummer', 'locatie']  # defineeer de te behouden velden
    fieldmappings = arcpy.FieldMappings()

    # stel de fieldmapping in
    fieldmappings.addTable(targetFeatures)
    fieldmappings.addTable(joinFeatures)
    keepers = velden

    # verwijder de niet-benodigde velden
    for field in fieldmappings.fields:
        if field.name not in keepers:
            fieldmappings.removeFieldMap(fieldmappings.findFieldMapIndex(field.name))
    # voer de spatial join uit
    arcpy.SpatialJoin_analysis(targetFeatures, joinFeatures, outfc, "#", "#", fieldmappings, match_option="INTERSECT",
                               search_radius=1)
    print "Rivierzijde profielen gekoppeld."

def voeg_velden_toe():
    # Set local variables
    inFeatures = "profielen_punten_z_join2"
    fieldName1 = "deltaX"
    fieldName2 = "deltaY"
    fieldName3 = "afstand"
    #fieldname4 = "omklap"
    fieldPrecision = 2


    # Execute AddField twice for two new fields
    arcpy.AddField_management(inFeatures, fieldName1, "FlOAT", fieldPrecision, field_is_nullable="NULLABLE")
    arcpy.AddField_management(inFeatures, fieldName2, "FlOAT", fieldPrecision, field_is_nullable="NULLABLE")
    arcpy.AddField_management(inFeatures, fieldName3, "FlOAT", fieldPrecision, field_is_nullable="NULLABLE")
    #arcpy.AddField_management(inFeatures, fieldname4, "FlOAT", fieldPrecision, field_is_nullable="NULLABLE")
    print "Velden toegevoegd."

def bereken_waardes():
    tbl =  'profielen_punten_z_join2'
    fields = ['x', 'y', 'type_lijn', 'deltaX', 'deltaY', 'profiel_id', 'rivierzijde','afstand']
    listp = []
    listx = []
    listy = []
    with arcpy.da.UpdateCursor(tbl, fields) as cur:
        for k, g in groupby(cur, lambda x: x[5]):

            for row in g:
                if row[2] == 0:
                    p = str(row[5])
                    x_vast = (row[0])
                    y_vast = (row[1])
                    listx.append(x_vast)
                    listy.append(y_vast)
                    listp.append(p)
                    #print row[5]
                else:
                    continue
    #print listp
    dictionary_X = dict(zip(listp, listx))
    dictionary_Y = dict(zip(listp, listy))




    with arcpy.da.UpdateCursor(tbl, fields) as cur3:
        for k, g in groupby(cur3, lambda x: x[5]):

            for row in g:
                nummer = str(row[5])
                x_0 = (dictionary_X[(nummer)])
                y_0 = (dictionary_Y[(nummer)])

                row[3] = row[0] - x_0
                if row[3] < 0:
                    row[3] = row[3] * -1
                else:
                  pass

                row[4] = row[1] - y_0
                if row[4] < 0:
                    row[4] = row[4] * -1
                else:
                    pass

                row[7] = math.sqrt((row[3]) ** 2 + (row[4]) ** 2)

                if row[7] < 0 and row[6] == 1:
                    row[7] = row[7]*-1
                elif row[7] > 0 and row[6] == None:
                    row[7] = row[7]*-1
                else:
                    pass

                # flip afstand voor Sweco
                row[7] = row[7]*-1


                cur3.updateRow(row)


    print "Waardes berekend."

# afmaken
def plotter():
    plt.style.use('seaborn-whitegrid')
    outpath = "C:/Users/vince/Desktop/plots_safe"
    # plt.style.use('seaborn-whitegrid')
    array = arcpy.da.FeatureClassToNumPyArray('profielen_punten_z_join2', ('profiel_id', 'dv_nummer', 'dijktraject', 'z_ahn', 'afstand'))
    df1 = pd.DataFrame(array)
    df = df1.dropna()
    grouped = df.groupby(["dv_nummer"])


    #dijktraject = '16-3'

    for name, group in grouped:

        dups = group.pivot_table(index=['profiel_id'], aggfunc='size')
        aantal_profielen = len(dups)


        color.cycle_cmap(aantal_profielen, cmap=plt.cm.binary)
        fig, ax = plt.subplots(figsize=(30, 5))

        for key, grp in group.groupby(['profiel_id']):
            x = grp['afstand']
            y = grp['z_ahn']


            ax.plot(x, y, linewidth=1, label= 'test')

        # plt.show()

        plt.savefig(path.join(outpath, "dijkvak_{0}.png".format(name)), pad_inches=0.02, dpi=300, bbox_inches='tight')
        plt.close()
        del fig, ax






# join_dv_indeling()

# generate_points()
# values_points()
# velden()
# koppel_centerline()
# knip_profielen()
# koppel_bovenlijn()
# voeg_velden_toe()
# bereken_waardes()
# plotter()


def create_df_xls(uitvoer_profielpunten):
    global d

    # import en dataframes
    array = arcpy.da.FeatureClassToNumPyArray(uitvoer_profielpunten, ('profiel_id', 'dv_nummer', 'dijktraject', 'z_ahn', 'afstand'))
    df_0 = pd.DataFrame(array)
    df_1 = df_0.dropna()
    df_round = df_1.round({'profiel_id': 0, 'dv_nummer': 0, 'z_ahn':2})
    sorted_df = df_round.sort_values(['dv_nummer'])

    # groupby
    grouped = sorted_df.groupby(["dv_nummer"])
    grouped.apply(lambda _df: _df.sort_values(by=['profiel_id']))

    # set afstanden index
    start = -150
    stop = 152
    step = 2
    index = list(range(start, stop, step))
    columns = ['afstand','z_ahn']

    # maak dataframes
    d={}

    aantal_groepen = []
    for name, item in grouped:
        aantal_groepen.append(item)
        name_int = int(name)
        d["dijkvak {}".format(name_int)] = pd.DataFrame(index=index)

# sort in groupby



    # vul dataframes
    for name, group in grouped:

        name_int = int(name)
        naam = "dijkvak {}".format(name_int)

        for key in d:
            if key == naam:
                ids_0 = group.profiel_id.unique().tolist()
                ids = sorted(ids_0)
                print ids



                df_group = d[naam]


                for item in ids:
                    list_z = []
                    list_afstand = []

                    for index, row in group.iterrows():
                        if row['profiel_id'] == item:

                            list_z.append(row['z_ahn'])
                            list_afstand.append(row['afstand'])


                    dictionary = dict(zip(list_afstand, list_z))

                    for index, row in df_group.iterrows():
                        for key in dictionary:
                            if key == index:
                                value = dictionary[key]
                                df_group.at[index,item] = value





def save_xls(dict_df, path):


    writer = ExcelWriter(excel)
    for key in sorted(d):
        d[key].to_excel(writer, key)

    writer.save()



def excel_aanpassen():
    array = arcpy.da.FeatureClassToNumPyArray(profielset, ('profiel_id', 'locatie'))
    df = pd.DataFrame(array)

    dictionary = dict(zip(df.profiel_id, df.locatie))

    wb = load_workbook(excel)

    for sheet in wb.worksheets:
        # sheet.insert_rows(1)

        # sheet['A1'] = ""
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                for key in dictionary:
                    if cell.value == key:
                        cell.value = dictionary[key]
                        #
                        #
                        # kolom = cell.column
                        # rij = 1
                        # # loc2 = get_column_letter(loc)
                        # # loc3 = loc2+str(1)
                        # # loc4 = str(loc3)
                        # # # sheet[loc4] = dictionary[key]
                        #
                        # # cell.value = dictionary[key]
                        # sheet.cell(row = rij, column = kolom).value = dictionary[key]
    wb.save(excel)



# save_xls(d,excel)
excel_aanpassen()